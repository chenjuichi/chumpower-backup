import os
import datetime
import pathlib
import csv
import time
import shutil
import psutil
import glob

import pymysql
from sqlalchemy import exc

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from database.tables import User, Session, Material, Bom, Assemble
from flask import Blueprint, jsonify, request, current_app

excelTable = Blueprint('excelTable', __name__)


# ------------------------------------------------------------------


#to convert date format
'''
def convert_date(date_str):
  try:
      return datetime.datetime.strptime(date_str, "%Y/%m/%d").date()
  except ValueError:
      return None
'''
def convert_date(date_value):
    try:
        return pd.to_datetime(date_value).date()  # 使用 pandas 轉換為日期
    except (ValueError, TypeError):
        return None


def close_open_file(filepath):
    # 遍歷所有進程
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # 檢查該進程打開的所有文件
            for item in proc.open_files():
                if item.path == filepath:
                    # 如果文件被打開，嘗試關閉
                    print(f"關閉進程 {proc.pid} 打開的文件: {filepath}")
                    proc.terminate()  # 終止進程
                    proc.wait()  # 等待進程終止
                    return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue
    return False


def pad_zeros(str):
    #length = len(str)
    #if length <= 1:
    return str.zfill(4)
    #else:
    #    return str.zfill(length + 3)


@excelTable.route("/fetchGlobalVar", methods=['GET'])
def fetch_global_var():
  print("fetchGlobalVar....")

  global global_var
  return jsonify({'value': global_var})


@excelTable.route("/countExcelFiles", methods=['GET'])
def count_excel_files():
    print("countExcelFiles....")

    _base_dir = current_app.config['baseDir']

    # 構建路徑模式，匹配以 "Report_" 開頭的 .xlsx 檔案
    path_pattern = f"{_base_dir}/Report_*.xlsx"
    # 使用 glob 找到所有符合條件的檔案
    files = glob.glob(path_pattern)
    # 計算檔案數量
    count = len(files)
    print("count:", count)
    # 如果沒有找到檔案，返回0
    #return count if count > 0 else 0
    return jsonify({
      'count': count
    })


@excelTable.route("/readAllExcelFiles", methods=['GET'])
def read_all_excel_files():
  print("readAllExcelFiles....")

  global global_var

  return_value = False
  return_message1 = '錯誤, 沒有工單檔案!'
  return_message2 = ''
  return_message = ''
  file_count_total = 0  #檔案總數目
  file_count_ok = 0     #檔案內總筆數

  code_to_assembleStep = {
    '109': 3,
    '106': 2,
    '110': 1
  }

  _base_dir = current_app.config['baseDir']
  _target_dir = _base_dir.replace("_in", "_out")
  print("read excel files, 目錄: ", _base_dir)
  print("move excel files to, 目錄: ", _target_dir)
  # 讀取指定目錄下的所有指定檔案名稱
  files = [f for f in os.listdir(_base_dir) if os.path.isfile(os.path.join(_base_dir, f)) and f.startswith('Report_') and f.endswith('.xlsx')]
  if (files):   #有工單檔案
    sheet_names_to_check = [
      current_app.config['excel_product_sheet'],
      current_app.config['excel_bom_sheet'],
      current_app.config['excel_work_time_sheet']
    ]
    _startRow = int(current_app.config['startRow'])

    s = Session()

    for _file_name in files:  #檔案讀取
      file_count_total +=1
      _path = _base_dir + '\\' + _file_name
      global_var = _path + ' 檔案讀取中...'

      #try:
      with open(_path, 'rb') as file:
        workbook = openpyxl.load_workbook(filename=file, read_only=True)
        return_value = True
        return_message1 = ''
        missing_sheets = [sheet for sheet in sheet_names_to_check if sheet not in workbook.sheetnames]

        if missing_sheets:
          return_value = False
          return_message1 = '錯誤, 工單檔案內沒有相關工作表!'
          print(return_message1)
          #continue                # 跳過當前檔案，繼續下一個檔案的處理
          break

        print(sheet_names_to_check[0] + ' sheet exists, data reading...')
        '''
        sheet = workbook[sheet_names_to_check[0]]             #取得工作表,組裝表
        _results = []
        for i in range(_startRow, sheet.max_row+1):
          if sheet.cell(row=i, column=1).value is None:   #資料讀取完畢
            break

          _order_num = str(sheet.cell(row = i, column = 2).value).strip()           #訂單編號
          _material_num = str(sheet.cell(row = i, column = 3).value).strip()        #料號(成品料號)
          _material_comment = str(sheet.cell(row = i, column = 4).value).strip()    #料號說明
          _material_qty = sheet.cell(row = i, column = 7).value                     #數量
          _material_date = sheet.cell(row = i, column = 5).value                    #建置日期

          newItem = s.query(Material).filter_by(order_num = _order_num, material_num=_material_num).first()
          return_value = True
          return_message2 =''
          if newItem:           #工單內容重複
            return_value = False
            return_message2 = '錯誤! 在' + _path + '檔案內, 系統已經有' + _order_num +'訂單及' + _material_num + '物料編號 相同資料...'
            print(return_message2)
            continue

          _obj = Material(
            order_num = _order_num,
            material_num = _material_num,
            material_comment = _material_comment,
            material_qty = _material_qty,
            material_date = _material_date,
          )

          _results.append(_obj)

          #file_count_ok += 1        #continue for loop

        s.bulk_save_objects(_results)

        s.commit()
        '''
        material_df = pd.read_excel(_path, sheet_name=0)  # First sheet for Material
        bom_df = pd.read_excel(_path, sheet_name=1)       # Second sheet for Bom
        assemble_df = pd.read_excel(_path, sheet_name=2)  # 3rd sheet for Assemble

        #print("material_df.columns:",material_df.columns)

        # Insert Material data
        for index, row in material_df.iloc[0:].iterrows():

          material = Material(
              order_num=row['單號'],
              material_num=row['料號'],
              material_comment=row['說明'],
              material_qty=row['數量'],
              material_date=convert_date(row['立單日']),
              material_delivery_date=convert_date(row['交期'])
          )
          s.add(material)
          s.commit()  # Commit each material individually to get the `id`

          material_order = str(row.iloc[1]).strip()                 #確保 row.iloc[1] 為 str 型別

          bom_df['訂單'] = bom_df['訂單'].fillna(0).astype(int)      #檢查是否存在 NaN 值
          bom_df['訂單'] = bom_df['訂單'].fillna('').astype(str)     #保持字串型態
          bom_entries = bom_df[bom_df.iloc[:, 0] == material_order] # 查詢對應的 BOM 項目
          print(f"bom_entries 中的資料筆數: {len(bom_entries)}")

          # Insert corresponding BOM entries
          for bom_index, bom_row in bom_entries.iterrows():

            bom = Bom(
                material_id=material.id,                  # Use the ID of the inserted material
                seq_num=bom_row['預留項目'],
                material_num=bom_row['物料'],
                material_comment=bom_row['物料說明'],
                req_qty=bom_row['需求數量'],
                start_date=convert_date(row['交期'])
            )
            s.add(bom)

            #print("Bom:", bom_index, bom_row.iloc[0])

          s.commit()

          #print(assemble_df.head())

          #assemble_entries = assemble_df[assemble_df.iloc[:, 0].astype(str) == str(material_order)]  #資料類型一致性
          assemble_entries = assemble_df[assemble_df.iloc[:, 0].astype(str).str.strip() == material_order.strip()]  #清除空格
          print(f"assemble_entries 中的資料筆數: {len(assemble_entries)}")
          #print("index:", index)

          # Insert corresponding Assemble entries
          for assemble_index, assemble_row in assemble_entries.iterrows():
            # 處理 NaN 值，將 NaN 替換為 None（SQLAlchemy 可以接受 None）
            reason = assemble_row['差異原因'] if not pd.isna(assemble_row['差異原因']) else None
            emp_num = assemble_row['員工號碼'] if not pd.isna(assemble_row['員工號碼']) else None
            confirm_comment = assemble_row['確認內文'] if not pd.isna(assemble_row['確認內文']) else None

            GMEIN = assemble_row['確認良品率 (GMEIN)']
            if GMEIN == 0:
              continue
            #
            workNum = assemble_row['工作中心']
            code = workNum[1:]             # 取得字串中的代碼 (去掉字串中的第一個字元)
            step_code = code_to_assembleStep.get(code, 0)   #
            #

            assemble = Assemble(
              material_id=material.id,                    # Use the ID of the inserted material
              material_num=assemble_row['物料'],
              material_comment=assemble_row['物料說明'],
              seq_num=assemble_row['作業'],
              work_num = workNum,
              process_step_code = step_code,
              meinh_qty=assemble_row['作業數量 (MEINH)'],
              good_qty = assemble_row['確認良品率 (GMEIN)'],           #確認良品數量
              non_good_qty = assemble_row['確認廢品 (MEINH)'],        #廢品數量

              reason = reason,
              user_id = emp_num,
              confirm_comment = confirm_comment,
            )
            s.add(assemble)

            #print("Assemble:", assemble_index, assemble_row.iloc[0])

          s.commit()



      #except pymysql.err.IntegrityError as e:
      #    s.rollback()
      #    print(f"IntegrityError: {e}")
      #except exc.IntegrityError as e:
      #    s.rollback()
      #    print(f"SQLAlchemy IntegrityError: {e}")
      #except Exception as e:
      #    s.rollback()
      #    print(f"Exception: {e}")





      # 移動檔案到目標目錄
      try:
        shutil.move(_path, _target_dir)
      except PermissionError as e:
        print(f"無法移動文件 {_path}，因為它仍然被佔用: {e}")

        #end for loop

    s.close()

  return jsonify({
    'status': return_value,
    'message': return_message1,
  })

@excelTable.route("/deleteAssemblesWithNegativeGoodQty", methods=['GET'])
def delete_assembles_with_negative_good_qty():
  print("deleteAssemblesWithNegativeGoodQty...")

  s = Session()

  # 查詢所有 Assemble 資料
  _objects = s.query(Assemble).all()

  # 用來追蹤前一筆資料
  previous_assemble = None

  # 迭代每一筆 Assemble 資料
  for current_assemble in _objects:
    if current_assemble.good_qty < 0:
      # 如果發現 ask_qty 小於 0，則刪除當前及前一筆 Assemble 紀錄
      if previous_assemble:
        print(f"Deleting previous assemble: {previous_assemble.id}, good_qty={previous_assemble.good_qty}")
        s.delete(previous_assemble)  # 刪除前一筆 Assemble 紀錄
      print(f"Deleting current assemble: {current_assemble.id}, ask_qty={current_assemble.good_qty}")
      s.delete(current_assemble)  # 刪除當前 Assemble 紀錄

      # 處理與 Material 的關聯（如有必要）
      # 根據業務邏輯決定是否刪除 Material，或更新 Material 的某些狀態
      material = s.query(Material).filter(Material.id == current_assemble.material_id).first()
      if material:
        # 更新 Material 的欄位，例如 isTakeOk 或其他狀態
        # material.isTakeOk = False  # 更新某些狀態
        # 如果需要刪除 Material，也可以使用：
        # s.delete(material)
        print(f"Updating or deleting related material: {material.id}, material_num={material.material_num}")

    # 更新 previous_assemble 為當前 Assemble
    previous_assemble = current_assemble

  # 提交刪除
  s.commit()
  s.close()

  return jsonify({
    'status': True,
    #'message': return_message1,
  })
