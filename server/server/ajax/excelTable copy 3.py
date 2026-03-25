import os
import datetime
import pathlib
import csv
import time
import shutil
import psutil
import glob
import math

import pymysql
from sqlalchemy import exc
from sqlalchemy import exists

import re

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from server.database.x_tables import User, Session, Material, Bom, Assemble, Product
from database.p_tables import P_Material, P_Bom, P_Assemble, P_Product, P_Part

from server.database.x_tables import ProcessedFile
from flask import Blueprint, jsonify, request, current_app

import warnings
warnings.filterwarnings(
   "ignore",
   category=FutureWarning,
  message=".*Setting an item of incompatible dtype.*"
)

excelTable = Blueprint('excelTable', __name__)

from log_util import setup_logger
logger = setup_logger(__name__)                 # 每個模組用自己的名稱


# ------------------------------------------------------------------


def read_all_excel_process_code_p():
    """
    從 _server_dir 目錄裡的 Excel 檔案（只處理 .xlsx/.xlsm），
    讀取工作表「配件工作中心資料表-0922 (2)」中
    欄位「製程代號 \n(標準內文碼)」的內容，

    若該欄位共有 m 筆有效資料：
      第 1 列 -> m
      第 2 列 -> m-1
      ...
      第 m 列 -> 1

    產生並回傳 code_to_assembleStep = { '100-01': m, '100-02': m-1, ... }
    （注意：這裡會把前面的 'B' 去掉，因為後面你有 code = workNum[1:]）
    """

    _base_dir = current_app.config['baseDir']
    _server_dir = _base_dir.replace("excel_in", "server")

    #print("read_all_excel_process_code_p(), _base_dir:", _base_dir)
    #print("read_all_excel_process_code_p(), _server_dir:", _server_dir)

    code_to_assembleStep = {}

    target_sheet_name = "配件工作中心資料表-0922 (2)"
    # 可能的欄位名稱
    target_col_candidates = [
      "製程代號 \n(標準內文碼)",
      "製程代號 (標準內文碼)",
    ]

    # 只掃描 openpyxl 支援的格式
    valid_exts = (".xlsx", ".xlsm", ".xltx", ".xltm")

    for root, dirs, files in os.walk(_server_dir):
        for fname in files:
            lower = fname.lower()

            # 1) 先過濾掉非 Excel 檔案（不再去開 .bat / .py）
            if not lower.endswith(valid_exts) and not lower.endswith(".xls"):
              continue

            full_path = os.path.join(root, fname)
            #print(f"處理 Excel 檔案: {full_path}")

            # 2) .xls 直接提示：openpyxl 不支援，要你改成 .xlsx
            if lower.endswith(".xls"):
              print(f"  檔案 {fname} 為舊版 .xls，openpyxl 不支援，請另存為 .xlsx 後再使用。")
              continue

            # 3) 其他副檔名（.xlsx/.xlsm…）才用 openpyxl 讀
            try:
              wb = load_workbook(full_path, data_only=True)
            except Exception as e:
              print(f"  載入工作簿失敗: {e}")
              continue

            if target_sheet_name not in wb.sheetnames:
              print(f"  檔案 {fname} 沒有工作表「{target_sheet_name}」，略過。")
              continue

            ws = wb[target_sheet_name]

            # 讀第一列當表頭
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
              print(f"  檔案 {fname} 的表頭是空的，略過。")
              continue

            # 找目標欄位的 column index（1-based）
            target_col_idx = None

            # 先直接比對候選名稱
            for idx, val in enumerate(header_row, start=1):
                if val is None:
                  continue
                val_str = str(val).strip()
                if val_str in target_col_candidates:
                  target_col_idx = idx
                  break

            # 找不到就模糊搜尋「製程代號」
            if target_col_idx is None:
                for idx, val in enumerate(header_row, start=1):
                  if val is None:
                    continue
                  val_str = str(val).strip()
                  if "製程代號" in val_str:
                    target_col_idx = idx
                    print(f"  發現相似欄位名稱使用: {val_str} (col {idx})")
                    break

            if target_col_idx is None:
                print(f"  檔案 {fname} 找不到「製程代號」欄位，略過。表頭: {list(header_row)}")
                continue

            # 取出該欄位從第2列開始的所有非空值
            codes = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if target_col_idx - 1 >= len(row):
                  continue
                cell_val = row[target_col_idx - 1]
                if cell_val is None:
                  continue
                cell_str = str(cell_val).strip()
                if cell_str == "":
                  continue

                # 統一把前面的 'B' 去掉，讓 key 變成 '100-01'
                # 這樣才能對上 read_all_excel_files_p 裡的 code = workNum[1:]
                if cell_str.startswith("B"):
                  cell_str = cell_str[1:]

                codes.append(cell_str)

            if not codes:
              print(f"  檔案 {fname} 的目標欄位沒有有效資料，略過。")
              continue

            m = len(codes)
            print(f"  檔案 {fname} 有 {m} 筆製程代號。")

            # 第1列 -> m, 第2列 -> m-1, ..., 第m列 -> 1
            for idx, code in enumerate(codes, start=1):
              value = m - idx + 1
              code_to_assembleStep[code] = value
              # print(f"    {idx}-th row, code={code}, value={value}")

    print("read_all_excel_process_code_p(), 完成，總筆數:", len(code_to_assembleStep))
    return code_to_assembleStep


# 安全轉整數的工具
def to_int0(v):
    if v is None:
        return 0
    s = str(v).strip()
    if s == '' or s.lower() == 'nan':
        return 0
    try:
        return int(float(s))
    except Exception:
        return 0


# 生成唯一檔案名稱的函式
def get_unique_filename(target_dir, filename, chip):
    base, ext = os.path.splitext(filename)  # 分離檔案名稱與副檔名
    counter = 1
    unique_filename = filename
    while os.path.exists(os.path.join(target_dir, unique_filename)):  # 檢查檔案是否已存在
        unique_filename = f"{base}_{chip}_{counter}{ext}"  # 為檔名新增後綴
        counter += 1
    return unique_filename


#to convert date format
def convert_date(date_value):
    try:
        return pd.to_datetime(date_value).date()  # 使用 pandas 轉換為日期
    except (ValueError, TypeError):
        return None


def close_open_file(filepath):
    # 遍歷所有進程
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # 檢查該進程打開的所有文件
            for item in proc.open_files():
                if item.path == filepath:
                    # 如果文件被打開，嘗試關閉
                    print(f"關閉進程 {proc.pid} 打開的文件: {filepath}")
                    proc.terminate()  # 終止進程
                    proc.wait()  # 等待進程終止
                    return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue
    return False


def pad_zeros(str):
  return str.zfill(4)


def clean_nan(value):
    """清理單一值的 NaN / NaT，轉為 None"""
    if pd.isna(value) or str(value).lower() == 'nan':
        return None
    return value


def normalize_order_number(value):
    """
    將單號轉為純字串，不帶小數點，並處理 NaN。
    例如：121100017702.0 -> '121100017702'
    """
    if pd.isna(value) or str(value).lower() == 'nan':
        return ''
    try:
        return str(int(float(value)))
    except (ValueError, TypeError):
        return str(value).strip()


@excelTable.route("/fetchGlobalVar", methods=['GET'])
def fetch_global_var():
  print("fetchGlobalVar....")

  global global_var
  return jsonify({'value': global_var})


@excelTable.route("/countExcelFilesP", methods=['GET'])
def count_excel_files_p():
    print("countExcelFilesP....")

    _base_dir = current_app.config['baseDir']
    _target_dir = _base_dir.replace("_in", "_out")
    _base_dir = _base_dir.replace("_in", "_in_p")
    print("read excel files, 目錄: ", _base_dir)
    print("move excel files to, 目錄: ", _target_dir)

    path_pattern = f"{_base_dir}/*.xlsx"

    files = glob.glob(path_pattern)         # 找到所有符合條件的檔案
    count = len(files)                      # 計算檔案數量
    print("count:", count)

    return jsonify({
      'count': count
    })


@excelTable.route("/countExcelFiles", methods=['GET'])
def count_excel_files():
    print("countExcelFiles....")

    _base_dir = current_app.config['baseDir']

    # 構建路徑模式，匹配以 "Report_" 開頭的 .xlsx 檔案
    #path_pattern = f"{_base_dir}/Report_*.xlsx"
    path_pattern = f"{_base_dir}/*.xlsx"
    # 使用 glob 找到所有符合條件的檔案
    files = glob.glob(path_pattern)
    # 計算檔案數量
    count = len(files)
    print("count:", count)
    # 如果沒有找到檔案，返回0
    #return count if count > 0 else 0
    return jsonify({
      'count': count
    })


@excelTable.route("/readAllExcelFilesP", methods=['GET'])
def read_all_excel_files_p():
  print("readAllExcelFilesP....")

  global global_var_p

  return_value = False
  return_message1 = '錯誤, 沒有工單檔案!'
  file_count_total = 0  #檔案總數

  #code_to_assembleStep = read_all_excel_process_code_p()
  #print("code_to_assembleStep:", code_to_assembleStep)

  _base_dir = current_app.config['baseDir']
  _target_dir = _base_dir.replace("_in", "_out")
  _base_dir = _base_dir.replace("_in", "_in_p")
  #print("read excel files, 目錄: ", _base_dir)
  #print("move excel files to, 目錄: ", _target_dir)

  # 讀取指定目錄下的所有指定檔案名稱
  files = [
     f for f in os.listdir(_base_dir)
     if os.path.isfile(os.path.join(_base_dir, f))
     and f.endswith('.xlsx')
  ]

  if not files:
    return jsonify({'status': False, 'message': '錯誤, 沒有工單檔案!'})

  sheet_names_to_check = [
    #current_app.config['excel_product_sheet'],
    current_app.config['p_excel_product_sheet'],
    current_app.config['excel_bom_sheet'],
    current_app.config['excel_work_time_sheet']
  ]

  s = Session()

  part_info_map = {}
  for p in s.query(P_Part).all():
      code = (p.part_code or '').strip()
      if not code:
          continue
      part_info_map[code] = {
          'comment': (p.part_comment or '').strip(),
          'process_step_code': int(p.process_step_code or 0)
      }

  for _file_name in files:  #檔案讀取, for loop_1
    file_count_total +=1
    file_name_base = re.sub(r'_copy_\d+', '', _file_name)   # 從檔名中移除像 _copy_1、_copy_2 這樣的字串，取得原始檔案名稱

    # 檢查是否已處理過
    already_processed = s.query(
      exists().where(ProcessedFile.file_name == file_name_base)
    ).scalar()

    if already_processed:
      return_message1 = f"檔案 {_file_name} 已處理過， 請再重整瀏覽器!"
      print(return_message1)
      _path = os.path.join(_base_dir, _file_name)
      file_count_total -=1
    else:
      _path = _base_dir + '\\' + _file_name
      global_var = _path + ' 檔案讀取中...'

      with open(_path, 'rb') as file:   # with loop_1_a
        workbook = openpyxl.load_workbook(filename=file, read_only=True)
        return_value = True
        return_message1 = ''

        print("workbook.sheetnames:", workbook.sheetnames)

        missing_sheets = [
          sheet for sheet in sheet_names_to_check if sheet not in workbook.sheetnames
        ]

        if missing_sheets:
          return jsonify({'status': False, 'message': '錯誤, 工單檔案內沒有相關工作表!'})

        print(sheet_names_to_check[0] + ' sheet exists, data reading...')

        material_df = pd.read_excel(_path, sheet_name=0)  # First sheet for Material
        bom_df = pd.read_excel(_path, sheet_name=1).fillna('')
        assemble_df = pd.read_excel(_path, sheet_name=2).fillna('')
        print("columns: material_df")
        print(list(material_df.columns))
        print("columns: bom_df")
        print(list(bom_df.columns))
        print("columns: assemble_df")
        print(list(assemble_df.columns))

        # 處理 BOM 和 Assemble 的 訂單欄位
        if '訂單' in bom_df.columns:
          bom_df.iloc[:, 0] = (
            bom_df.iloc[:, 0]
            .apply(normalize_order_number)
            .replace('nan', '')
            .astype(str)
          )
          bom_df = bom_df.astype({bom_df.columns[0]: object})   # 把 dtype 設成 object
        else:
          # 如果沒資料，建立一個空的 DataFrame，或直接跳過
          bom_df = pd.DataFrame(columns=["訂單"])  # 預設空結構
          print("⚠️ bom_df 沒有資料或缺少 '訂單' 欄位，已建立空 DataFrame")
        # end if

        print("bom_df:", bom_df, len(bom_df) )

        if '訂單' in assemble_df.columns:
          assemble_df['訂單'] = (
            assemble_df['訂單']
            .apply(normalize_order_number)
            .replace('nan', '')
            .astype(str)
          )
          assemble_df = assemble_df.astype({'訂單': object})
        else:
          # 如果沒有訂單欄位，就印出警告，後面就一定抓不到資料
          print("⚠️ assemble_df 缺少 '訂單' 欄位，無法依訂單過濾工序資料! columns =", list(assemble_df.columns))

        # 先記錄整個 BOM sheet 是否完全沒資料
        bom_df_is_empty = bom_df.empty

        # 記錄已處理過的單號，避免重複產生 P_Material / P_Product / P_Bom / P_Assemble
        seen_order_nums = set()

        # 處理 Material
        for _, row in material_df.iterrows(): # for loop_material
          order_num = normalize_order_number(row.get('單號'))
          if not order_num:
            print(f"[警告] 單號為空，跳過該筆資料: {row.to_dict()}")
            continue

          # 🔹若這個單號之前處理過，就略過
          if order_num in seen_order_nums:
            print(f"[略過] 單號 {order_num} 已在前面處理過，略過重複的 Material/BOM/Assemble")
            continue

          seen_order_nums.add(order_num)

          # 先檢查 BOM 是否有資料
          if bom_df_is_empty:
              # 整個 BOM sheet 是空的情況：
              # 👉 不要略過，後面會幫這個單號建立一筆「預設」的 P_Bom
              bom_entries = None
              print(f"[預設] 整體 BOM 為空，單號 {order_num} 仍建立 Material / Product / Assemble，並建 1 筆預設 P_Bom。")
          else:
            bom_entries = bom_df[bom_df['訂單'] == order_num]
            if bom_entries.empty:
                print(f"[略過] 單號 {order_num} 沒有 BOM 資料，不建立 Material")
                continue   # 直接跳過，不建 Material / Product / Assemble

          # data
          # --------- 有 BOM 才建立 Material ----------
          tempQty = clean_nan(row.get('數量')) or 0
          temp_sd_time_B100 = clean_nan(row.get('B100加工(一)工時(分)')) or 0
          temp_sd_time_B102 = clean_nan(row.get('B102KL加工綜合工時(分)')) or 0
          temp_sd_time_B103 = clean_nan(row.get('B103KT加工工時(分)')) or 0
          temp_sd_time_B107 = clean_nan(row.get('B107震動研磨工時(分)')) or 0
          temp_sd_time_B108 = clean_nan(row.get('B108綜合加工工時(分)')) or 0

          material_isBom = bom_df_is_empty
          material_isTakeOk = bom_df_is_empty
          material_isShow = bom_df_is_empty

          material = P_Material(
            order_num=order_num,
            material_num=clean_nan(row.get('料號')) or '',
            material_comment=clean_nan(row.get('說明')) or '',
            material_qty=tempQty,
            material_date=convert_date(row.get('立單日')),
            material_delivery_date=convert_date(row.get('交期')),
            total_delivery_qty=tempQty,
            sd_time_B100="{:.2f}".format(float(temp_sd_time_B100)),
            sd_time_B102="{:.2f}".format(float(temp_sd_time_B102)),
            sd_time_B103="{:.2f}".format(float(temp_sd_time_B103)),
            sd_time_B107="{:.2f}".format(float(temp_sd_time_B107)),
            sd_time_B108="{:.2f}".format(float(temp_sd_time_B108)),
            move_by_automatic_or_manual = False,
            move_by_process_type = 4,
            isBom = material_isBom,
            isTakeOk = material_isTakeOk,
            isShow = material_isShow,
            show1_ok = '2' if material_isBom else '1',
            show2_ok = '3' if material_isBom else '0',
            delivery_qty = tempQty if material_isBom else 0,
          )
          s.add(material)
          s.flush()  # 確保 material.id 可用

          # Product
          product = P_Product(material_id=material.id)
          s.add(product)
          s.commit()

          # BOM
          if not bom_df_is_empty: # bom_df_is_empty if block
            # 一般情況：有 BOM 資料，就按原本邏輯依據訂單編號建立多筆 P_Bom
            bom_entries = bom_df[bom_df['訂單'] == order_num]
            print(f"bom_entries 中的資料筆數: {len(bom_entries)}")

            for _, bom_row in bom_entries.iterrows():  # for loop_bom
                shortage = to_int0(bom_row.get('物料短缺'))
                bom = P_Bom(
                    material_id=material.id,
                    seq_num=clean_nan(bom_row.get('預留項目')),
                    material_num=clean_nan(bom_row.get('物料')),
                    material_comment=clean_nan(bom_row.get('物料說明')),
                    req_qty=clean_nan(bom_row.get('需求數量')),
                    start_date=convert_date(row.get('交期')),
                    lack_bom_qty=shortage,
                    receive=(shortage == 0),
                )
                s.add(bom)
            # end for loop_bom
          else:
            # ⚠️ 特例：整個 BOM sheet 是空的 (bom_df: Empty DataFrame)
            # 為每一個 Material 建立一筆「預設」的 P_Bom
            print(f"[預設 BOM] 單號 {order_num} BOM 為空，建立 1 筆預設 P_Bom")
            bom = P_Bom(
                material_id=material.id,
                seq_num='',
                material_num='',
                material_comment='預設BOM (原始Excel無BOM資料)',
                req_qty=0,   # 這裡如果你希望等於 tempQty 也可以改成 tempQty
                start_date=convert_date(row.get('交期')),
                lack_bom_qty=0,
                receive=True,
            )
            s.add(bom)
          # end bom_df_is_empty if block

          s.commit()

          # Assemble
          if '訂單' not in assemble_df.columns:
            print(f"⚠️ assemble_df 沒有 '訂單' 欄位，無法產生該工單的 P_Assemble，order_num={order_num}")
            assemble_entries = assemble_df.iloc[0:0]  # 空的 DataFrame
          else:
            assemble_entries = assemble_df[assemble_df['訂單'] == order_num]

          print(f"assemble_entries 中的資料筆數: {len(assemble_entries)}")

          # 只拿「訂單 + 物料」都跟這個 material 相同的工序
          material_num = material.material_num
          group_df = assemble_entries[assemble_entries['物料'] == material_num].reset_index(drop=True)
          n = len(group_df)

          # 預設全部 isSimultaneously = False
          simultaneously_flags = [False] * n

          # 先掃一次：找出「作業有值但標準內文碼為空」的列，標記它的前一筆 / 下一筆
          for i in range(n):
            seq_val = clean_nan(group_df.loc[i, '作業'])
            seq_str = str(seq_val).strip() if seq_val is not None else ''

            code_val = clean_nan(group_df.loc[i, '標準內文碼'])
            code_str = str(code_val).strip() if code_val is not None else ''

            # 符合「有作業、沒標準內文碼」的條件
            if seq_str and not code_str:
              # 上一筆
              if i - 1 >= 0:
                # 同一個物料，才算同組（其實 group_df 已經是同物料了）
                simultaneously_flags[i - 1] = True
              # 下一筆
              if i + 1 < n:
                simultaneously_flags[i + 1] = True

          processed_work_nums = set()

          # 再掃一次：真正建立 P_Assemble
          for i in range(n):
            assemble_row = group_df.loc[i]

            workNum = clean_nan(assemble_row.get('標準內文碼'))
            # 空 workNum 的那筆（例如中間那筆沒有標準內文碼）本來就不建 P_Assemble
            if not workNum:
              print("    ⚠️ workNum 為空，略過這筆工序。")
              continue

            if workNum in processed_work_nums:
              print(f"    ⚠️ workNum {workNum} 重複，已處理過，略過。")
              continue
            processed_work_nums.add(workNum)

            seq_num = clean_nan(assemble_row.get('作業'))

            # 取 code & step_code
            #if len(workNum) > 1:
            #  code = workNum[1:]
            #else:
            #  code = workNum  # 安全一點，避免 workNum 只有 1 個字的時候 [1:] 變成空字串

            #step_code = code_to_assembleStep.get(code, 0)
            part_info = part_info_map.get(workNum)
            #print("part_info_map", part_info_map)
            #print("part_info_map, workNum", workNum)
            #print("part_info_map, part_info", part_info)
            step_code = part_info['process_step_code']

            #print(f"    ▶ code: {repr(code)}, step_code: {step_code}")
            print(f"    ▶ step_code: {step_code}")

            abnormal_field = False

            # 讀取作業短文，判斷是否以 'Z' 開頭 → isStockIn
            op_short = clean_nan(assemble_row.get('作業短文')) or ''
            op_short_str = str(op_short).strip()
            is_stock_in = op_short_str.startswith('Z')

            # 這一列的 isSimultaneously
            is_simultaneously = simultaneously_flags[i]

            #if bom_df_is_empty:
            #  must_receive_qty=clean_nan(row.get('數量')) or 0
            #else:
            #  must_receive_qty=0

            assemble = P_Assemble(
                material_id=material.id,
                material_num=clean_nan(assemble_row.get('物料')),
                material_comment=clean_nan(assemble_row.get('物料說明')),
                seq_num=seq_num,
                work_num=workNum,
                process_step_code=step_code,
                input_abnormal_disable=abnormal_field,
                user_id='',
                isStockIn=is_stock_in,
                isSimultaneously=is_simultaneously,
                #must_receive_qty=must_receive_qty,
                must_receive_qty=clean_nan(assemble_row.get('作業數量 (MEINH)')),
                isShowBomGif = material_isBom,
            )
            s.add(assemble)

          # end loop_assemble
          s.commit()

          # commit 後再查一次看這個 material_id 到底有幾筆
          cnt = s.query(P_Assemble).filter_by(material_id = material.id).count()
          print(f"  [P_Assemble] material_id={material.id} 在 P_Assemble 目前總共有 {cnt} 筆")

    # 移動處理完成的檔案
    try:
      unique_filename = get_unique_filename(_target_dir, _file_name, "copy")
      unique_target_path = os.path.join(_target_dir, unique_filename)
      print("unique_target_path:", unique_target_path)
      shutil.move(_path, unique_target_path)
      print(f"檔案 {_path} 已成功移動到 {unique_target_path}")
    except Exception as e:
      print(f"移動檔案時發生錯誤: {e}")

    continue

  #end for loop_1
  s.close()

  return jsonify({
    'status': return_value,
    'message': return_message1,
  })


@excelTable.route("/readAllExcelFiles", methods=['GET'])
def read_all_excel_files():
  print("readAllExcelFiles....")

  global global_var

  return_value = False
  return_message1 = '錯誤, 沒有工單檔案!'
  file_count_total = 0  #檔案總數

  # 2025-06-12, 改順序
  code_to_assembleStep = {    #組裝區工作順序, 3:最優先
    '109': 3,
    #'106': 2, '110': 1,
    '106': 1, '110': 2,
  }

  _base_dir = current_app.config['baseDir']
  _target_dir = _base_dir.replace("_in", "_out")
  print("read excel files, 目錄: ", _base_dir)
  print("move excel files to, 目錄: ", _target_dir)

  ## 取得 excel_out 內已處理檔案（移除 _copy_x 尾碼以便比對）
  #processed_files = {
  #    re.sub(r'_copy_\d+', '', f)
  #    for f in os.listdir(_target_dir)
  #    if os.path.isfile(os.path.join(_target_dir, f)) and f.endswith('.xlsx')
  #}
  #

  # 讀取指定目錄下的所有指定檔案名稱
  #files = [f for f in os.listdir(_base_dir) if os.path.isfile(os.path.join(_base_dir, f)) and f.startswith('Report_') and f.endswith('.xlsx')]
  #files = [f for f in os.listdir(_base_dir) if os.path.isfile(os.path.join(_base_dir, f)) and f.endswith('.xlsx')]
  files = [
     f for f in os.listdir(_base_dir)
     if os.path.isfile(os.path.join(_base_dir, f))
     and f.endswith('.xlsx')
  ]

  '''
  # 排除已處理的檔案
  files = [
      f for f in os.listdir(_base_dir)
      if os.path.isfile(os.path.join(_base_dir, f))
        and f.endswith('.xlsx')
        and f not in processed_files
  ]
  '''
  if not files:
    return jsonify({'status': False, 'message': return_message1})

  #if (files):   #有工單檔案, if condition_a
  sheet_names_to_check = [
    current_app.config['excel_product_sheet'],
    current_app.config['excel_bom_sheet'],
    current_app.config['excel_work_time_sheet']
  ]
  #_startRow = int(current_app.config['startRow'])

  s = Session()

  for _file_name in files:  #檔案讀取, for loop_1
    file_count_total +=1
    file_name_base = re.sub(r'_copy_\d+', '', _file_name)   # 從檔名中移除像 _copy_1、_copy_2 這樣的字串，取得原始檔案名稱

    # 檢查是否已處理過
    already_processed = s.query(
      exists().where(ProcessedFile.file_name == file_name_base)
    ).scalar()

    if already_processed:
      return_message1 = f"檔案 {_file_name} 已處理過， 請再重整瀏覽器!"
      print(return_message1)
      _path = os.path.join(_base_dir, _file_name)
      file_count_total -=1
    else:
      _path = _base_dir + '\\' + _file_name
      global_var = _path + ' 檔案讀取中...'

      with open(_path, 'rb') as file:   # with loop_1_a
        workbook = openpyxl.load_workbook(filename=file, read_only=True)
        return_value = True
        return_message1 = ''

        #missing_sheets = [sheet for sheet in sheet_names_to_check if sheet not in workbook.sheetnames]
        missing_sheets = [
          sheet for sheet in sheet_names_to_check if sheet not in workbook.sheetnames
        ]

        if missing_sheets:
          return jsonify({'status': False, 'message': '錯誤, 工單檔案內沒有相關工作表!'})

          #return_value = False
          #return_message1 = '錯誤, 工單檔案內沒有相關工作表!'
          #print(return_message1)
          #break

        print(sheet_names_to_check[0] + ' sheet exists, data reading...')

        material_df = pd.read_excel(_path, sheet_name=0)  # First sheet for Material
        # 2025-09-09 add
        # 欄位級正規化
        if '單號' in material_df.columns:
          material_df['單號'] = (
              material_df['單號']
              .apply(normalize_order_number)
              .astype(str)
              .replace('nan', '')
          )
        #
        #bom_df = pd.read_excel(_path, sheet_name=1).fillna('')
        #
        bom_df = pd.read_excel(_path, sheet_name=1)
        # 文本欄位可補空字串
        for col in ['物料', '物料說明']:
          if col in bom_df.columns:
            bom_df[col] = bom_df[col].fillna('')
        # 數值欄位轉數字
        for col in ['物料短缺', '需求數量', '預留項目']:
          if col in bom_df.columns:
            bom_df[col] = pd.to_numeric(bom_df[col], errors='coerce').fillna(0).astype(int)
        #
        assemble_df = pd.read_excel(_path, sheet_name=2).fillna('')

        # 統一處理 BOM 和 Assemble 的 訂單欄位
        if '訂單' in bom_df.columns:
          # 2025-08-11 modify
          #bom_df['訂單'] = bom_df['訂單'].apply(normalize_order_number)
          '''
          bom_df['訂單'] = (
            bom_df['訂單']
            .apply(normalize_order_number)
            .astype(str)
            .replace('nan', '')
          )
          '''
          bom_df.iloc[:, 0] = (
              bom_df.iloc[:, 0]
              .apply(normalize_order_number)
              .replace('nan', '')
              .astype(str)   # 直接最後轉成 str → object
          )

          # 明確把 dtype 設成 object
          bom_df = bom_df.astype({bom_df.columns[0]: object})

        # 2025-08-11 modify
        #assemble_df.iloc[:, 0] = assemble_df.iloc[:, 0].apply(normalize_order_number)
        '''
        assemble_df.iloc[:, 0] = (
          assemble_df.iloc[:, 0]
          .apply(normalize_order_number)
          .astype(str)
          .replace('nan', '')
        )
        '''
        assemble_df.iloc[:, 0] = (
            assemble_df.iloc[:, 0]
            .apply(normalize_order_number)
            .replace('nan', '')
            .astype(str)
        )
        assemble_df = assemble_df.astype({assemble_df.columns[0]: object})

        # 處理 Material
        for _, row in material_df.iterrows(): # for loop_material
            order_num = normalize_order_number(row.get('單號'))
            if not order_num:
                print(f"[警告] 單號為空，跳過該筆資料: {row.to_dict()}")
                continue

            tempQty = clean_nan(row.get('數量')) or 0
            temp_sd_time_B109 = clean_nan(row.get('B109組裝工時(分)')) or 0
            temp_sd_time_B106 = clean_nan(row.get('B106雷刻工時(分)')) or 0
            temp_sd_time_B110 = clean_nan(row.get('B110檢驗工時(分)')) or 0

            material = Material(
                order_num=order_num,
                material_num=clean_nan(row.get('料號')) or '',
                material_comment=clean_nan(row.get('說明')) or '',
                material_qty=tempQty,
                material_date=convert_date(row.get('立單日')),
                material_delivery_date=convert_date(row.get('交期')),
                total_delivery_qty=tempQty,
                sd_time_B109="{:.2f}".format(float(temp_sd_time_B109)),
                sd_time_B106="{:.2f}".format(float(temp_sd_time_B106)),
                sd_time_B110="{:.2f}".format(float(temp_sd_time_B110)),
            )
            s.add(material)
            s.flush()  # 確保 material.id 可用

            # Product
            product = Product(material_id=material.id)
            s.add(product)
            s.commit()

            # BOM
            bom_entries = bom_df[bom_df['訂單'] == order_num]
            print(f"bom_entries 中的資料筆數: {len(bom_entries)}")

            for _, bom_row in bom_entries.iterrows(): # for loop_bom
                #temp=clean_nan(bom_row.get('物料短缺'))
                shortage = to_int0(bom_row.get('物料短缺'))
                bom = Bom(
                    material_id=material.id,
                    seq_num=clean_nan(bom_row.get('預留項目')),
                    material_num=clean_nan(bom_row.get('物料')),
                    material_comment=clean_nan(bom_row.get('物料說明')),
                    req_qty=clean_nan(bom_row.get('需求數量')),
                    start_date=convert_date(row.get('交期')),
                    #lack_bom_qty=temp,
                    lack_bom_qty = shortage,
                    #receive= True if temp==0 else False,
                    receive=(shortage == 0),
                )
                s.add(bom)
            s.commit()

            # Assemble
            assemble_entries = assemble_df[assemble_df.iloc[:, 0] == order_num]
            print(f"assemble_entries 中的資料筆數: {len(assemble_entries)}")

            processed_work_nums = set()
            for _, assemble_row in assemble_entries.iterrows(): # for loop_assemble
                workNum = clean_nan(assemble_row.get('工作中心'))
                if not workNum or workNum in processed_work_nums:
                    continue
                processed_work_nums.add(workNum)

                code = workNum[1:]
                step_code = code_to_assembleStep.get(code, 0)
                #abnormal_field = (workNum == 'B109')     # 2025-07-29 modify
                abnormal_field = False

                assemble = Assemble(
                    material_id=material.id,
                    material_num=clean_nan(assemble_row.get('物料')),
                    material_comment=clean_nan(assemble_row.get('物料說明')),
                    seq_num=clean_nan(assemble_row.get('作業')),
                    work_num=workNum,
                    process_step_code=step_code,
                    input_abnormal_disable=abnormal_field,
                    user_id=''
                )
                s.add(assemble)
            s.commit()
        # end for loop_material

        # ✅ 資料處理完後，記錄檔案處理紀錄
        processed_file = ProcessedFile(file_name=file_name_base)
        s.add(processed_file)
        s.commit()

    # 移動處理完成的檔案
    try:
      unique_filename = get_unique_filename(_target_dir, _file_name, "copy")
      unique_target_path = os.path.join(_target_dir, unique_filename)
      print("unique_target_path:", unique_target_path)
      shutil.move(_path, unique_target_path)
      print(f"檔案 {_path} 已成功移動到 {unique_target_path}")
    except Exception as e:
      print(f"移動檔案時發生錯誤: {e}")

    continue

  #end for loop_1
  s.close()
  #end if condition_a

  return jsonify({
    'status': return_value,
    'message': return_message1,
  })


@excelTable.route("/deleteAssemblesWithNegativeGoodQty", methods=['GET'])
def delete_assembles_with_negative_good_qty():
  print("deleteAssemblesWithNegativeGoodQty...")

  s = Session()

  # 查詢所有 Assemble 資料
  _objects = s.query(Assemble).all()

  # 用來追蹤前一筆資料
  previous_assemble = None

  # 迭代每一筆 Assemble 資料
  for current_assemble in _objects:
    if current_assemble.good_qty < 0:
      # 如果發現 ask_qty 小於 0，則刪除當前及前一筆 Assemble 紀錄
      if previous_assemble:
        print(f"Deleting previous assemble: {previous_assemble.id}, good_qty={previous_assemble.good_qty}")
        s.delete(previous_assemble)  # 刪除前一筆 Assemble 紀錄
      print(f"Deleting current assemble: {current_assemble.id}, ask_qty={current_assemble.good_qty}")
      s.delete(current_assemble)  # 刪除當前 Assemble 紀錄

      # 處理與 Material 的關聯（如有必要）
      # 根據業務邏輯決定是否刪除 Material，或更新 Material 的某些狀態
      material = s.query(Material).filter(Material.id == current_assemble.material_id).first()
      if material:
        # 更新 Material 的欄位，例如 isTakeOk 或其他狀態
        # material.isTakeOk = False  # 更新某些狀態
        # 如果需要刪除 Material，也可以使用：
        # s.delete(material)
        print(f"Updating or deleting related material: {material.id}, material_num={material.material_num}")

    # 更新 previous_assemble 為當前 Assemble
    previous_assemble = current_assemble

  # 提交刪除
  s.commit()
  s.close()

  return jsonify({
    'status': True,
    #'message': return_message1,
  })


@excelTable.route("/exportToExcelForError", methods=['POST'])
def export_to_excel_for_error():
    print("exportToExcelForError....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']

    print("_blocks:", _blocks)

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    file_name = '組裝異常記錄查詢_'+today + '.xlsx'
    current_file = 'C:\\vue\\chumpower\\excel_export\\'+ file_name

    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功

    if file_check:
      try:
        os.remove(current_file)  # 刪除舊檔案
      except PermissionError:
        print(f"無法刪除 {current_file}，請確認是否已關閉該檔案")
        return_value = False
        return jsonify({"success": False, "message": "請關閉 Excel 檔案後重試"})
    #if file_check:
    #  wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
    #else:
    #  wb = Workbook()     # 建立空白的 Excel 活頁簿物件
    wb = Workbook()     # 建立空白的 Excel 活頁簿物件
    # ws = wb.active
    ws = wb.worksheets[0]   # 取得第一個工作表

    ws.title = '組裝異常記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
    ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

    for obj in _blocks:
      temp_array = []

      # 在寫入 Excel 時做轉換
      raw_str = obj['cause_message_str']  # e.g., "混料(M01001),散爪(M01002),掉爪(M01003)"
      messages = [re.sub(r'\(.*?\)', '', m).strip() for m in raw_str.split(',')]  # 移除括號和裡面內容
      cleaned_str = ','.join(messages)  # "混料,散爪,掉爪"

      temp_array.append(obj['order_num'])
      temp_array.append(obj['comment'])
      temp_array.append(obj['delivery_date'])
      temp_array.append(obj['req_qty'])
      temp_array.append(obj['delivery_qty'])
      temp_array.append(obj['user'])
      #temp_array.append(obj['cause_message_str'])
      temp_array.append(cleaned_str)
      temp_array.append(obj['cause_user'])
      temp_array.append(obj['cause_date'])

      ws.append(temp_array)

    for col in ws.columns:
      column = col[0].column_letter  # Get the column name
      temp_cell = column + '1'
      ws[temp_cell].font = Font(
          name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
      ws[temp_cell].alignment = Alignment(horizontal='center')
      ws.column_dimensions[column].bestFit = True

    #wb.save(current_file)
    #預防 Excel 檔案被鎖定
    temp_file = current_file.replace(".xlsx", "_temp.xlsx")
    wb.save(temp_file)
    os.replace(temp_file, current_file)  # 用新檔案取代舊檔案

    print("file_name:", file_name)

    return jsonify({
      'status': return_value,
      'message': current_file,
      'file_name': file_name,
      # 'outputs': myDir,
    })


@excelTable.route("/exportToExcelForAssembleInformation", methods=['POST'])
def export_to_excel_for_assemble_information():
    import os
    import datetime
    from flask import jsonify, request
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    print("hello, exportToExcelForAssembleInformation....")

    request_data = request.get_json(force=True) or {}
    _blocks = request_data.get('blocks', [])
    _name = request_data.get('name', '')

    now = datetime.datetime.now()
    today = now.strftime('%Y-%m-%d-%H%M')
    file_name = f'組裝區在製品生產資訊查詢_{today}.xlsx'
    export_dir = r'C:\vue\chumpower\excel_export'
    os.makedirs(export_dir, exist_ok=True)
    current_file = os.path.join(export_dir, file_name)

    code_to_name = {
        1 : '備料',
        19: '等待AGV(備料區)',
        2:  'AGV運行(備料區->組裝區)',
        20: 'AGV運行到組裝區',
        21: '組裝',
        22: '檢驗',
        23: '雷射',
        29: '等待AGV(組裝區)',
        3:  'AGV運行(組裝區->成品區)',
        30: 'AGV運行到成品區',
        31: '成品入庫',
        5:  '堆高機運行(備料區->組裝區)',
        6:  '堆高機運行(組裝區->成品區)',
    }

    # === 1) 安全時間轉換：吃 str/datetime/date/None，回 datetime 或 None ===
    def to_dt(val):
        if not val:
            return None
        if isinstance(val, datetime.datetime):
            return val
        if isinstance(val, datetime.date):
            return datetime.datetime.combine(val, datetime.time.min)
        if isinstance(val, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                        "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
                try:
                    return datetime.datetime.strptime(val, fmt)
                except ValueError:
                    continue
        return None

    # === 2) 若舊檔存在，嘗試刪除；占用時回報請關閉 ===
    if os.path.exists(current_file):
        try:
            os.remove(current_file)
        except PermissionError:
            return jsonify({"status": False, "message": "請關閉 Excel 檔案後重試", "file_name": ""}), 200

    # === 3) 建簿/表頭 ===
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = '組裝區在製品生產資訊查詢-' + _name
    ws.sheet_properties.tabColor = '7da797'

    header = ['訂單編號', '說明', '交期', '訂單數量', '現況數量', '工序', '開始時間', '結束時間']
    ws.append(header)
    for idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(name='微軟正黑體', color='FF0000', bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 16

    s = Session()
    temp_file = current_file.replace(".xlsx", "_temp.xlsx")

    try:
        for obj in _blocks:
            order_num = obj.get('order_num', '')
            # 先寫 Material 主列
            ws.append([
                order_num,
                obj.get('comment', ''),
                obj.get('delivery_date', ''),
                obj.get('req_qty', ''),
                obj.get('delivery_qty', ''),
                '', '', ''  # 讓出工序/開始/結束欄位
            ])

            material = s.query(Material).filter(Material.order_num == order_num).first()

            # 取總量（可能 None）
            work_qty = (material.total_delivery_qty if material and getattr(material, "total_delivery_qty", None) else 0)

            processes = material._process if (material and getattr(material, "_process", None)) else []
            for process in processes:
                ptype = getattr(process, "process_type", None)
                status = code_to_name.get(ptype, '空白')

                # === 4) 員工資訊（可能沒有）===
                emp_id = (getattr(process, "user_id", "") or "")
                if emp_id:
                    user = s.query(User).filter_by(emp_id=emp_id).first()
                    emp_short = emp_id.lstrip("0") if isinstance(emp_id, str) else str(emp_id)
                    if user and getattr(user, "emp_name", None):
                        status = f"{status}({emp_short}{user.emp_name})"

                # === 5) 安全時間解析 + 進行中處理 ===
                start_dt = to_dt(getattr(process, "begin_time", None))
                end_dt = to_dt(getattr(process, "end_time", None)) if ptype != 31 else None

                # 進行中：end_dt 為 None → 你可以選：A 顯示空白（這版用空白），B 用 now 當暫止
                effective_end = end_dt if end_dt is not None else (now if start_dt else None)

                # 區段分鐘（不一定要寫入；給你未來算單件工時用）
                period_minutes = None
                if start_dt and effective_end:
                    td = effective_end - start_dt
                    if td.total_seconds() >= 0:
                        period_minutes = int(td.total_seconds() // 60)

                # === 6) 輸出到 Excel（時間欄位做安全格式化）===
                def fmt(dt):
                    return dt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(dt, datetime.datetime) else ''

                ws.append([
                    '', '', '', '', '',
                    status,
                    fmt(start_dt),
                    fmt(end_dt),   # 想顯示「目前時間」就改成 fmt(effective_end)
                ])

        # === 7) 先存暫存檔再原子替換 ===
        wb.save(temp_file)
        os.replace(temp_file, current_file)

        print("file_name:", file_name)
        return jsonify({'status': True, 'message': current_file, 'file_name': file_name}), 200

    except Exception as e:
        s.rollback()
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception:
            pass
        print("export_to_excel_for_assemble_information EXCEPTION:", repr(e))
        return jsonify({'status': False, 'message': f'匯出失敗: {e}', 'file_name': ''}), 500
    finally:
        s.close()


# 上傳.xlsx工單檔案的 API
@excelTable.route('/uploadExcelFile', methods=['POST'])
def upload_excel_file():
  print("uploadExcelFile....")

  _base_dir = current_app.config['baseDir']

  # 以 baseDir 的上層作為根目錄（通常 baseDir 會是 .../excel_in）
  base_dir_default = os.path.abspath(current_app.config['baseDir'])
  root_dir = os.path.abspath(os.path.dirname(base_dir_default))

  if not os.path.exists(_base_dir):
    os.makedirs(_base_dir)

  file = request.files.get('file')

  _upload_type = request.form.get('uploadType')
  upload_type=_upload_type.strip()
  print("uploadType:", upload_type)

  if not file:
    return jsonify({'message': '沒有選擇檔案'}), 400

  # 確保是 Excel 檔案
  if not file.filename.endswith(('.xlsx', '.xls')):
    return jsonify({'message': '只允許上傳 Excel 檔案'}), 400

  if file.filename == '':
    return jsonify({'message': '沒有選擇檔案'}), 400

  if upload_type == "excelp":
    print("step1...")
    _base_dir = os.path.join(os.path.dirname(_base_dir), "excel_in_p")
    print("base_dir:", _base_dir)

  if upload_type == "excelm":
    print("step2...")
    _base_dir = os.path.join(os.path.dirname(_base_dir), "excel_modify")
    print("base_dir:", _base_dir)

  #

  #
  file_path = os.path.join(_base_dir, file.filename)
  print("file_path:",file_path)

  # 檢查檔案是否已存在
  if os.path.exists(file_path):
    #return jsonify({'message': f'檔案 "{file.filename}" 已存在'}), 400
    unique_filename = get_unique_filename(_base_dir, file.filename, "copy")
    file_path = os.path.join(_base_dir, unique_filename)

  file.save(file_path)

  try:
    with open(file_path, 'rb') as f:
        print(f"檔案成功儲存並可讀取: {file_path}")
        return jsonify({
          'message': '上傳成功',
          'filename': file.filename,
          'status': True,
        })
  except Exception as e:
      print(f"檔案儲存後讀取失敗: {str(e)}")
      return jsonify({'message': f'檔案儲存後讀取失敗: {str(e)}'}), 500
  #return jsonify({'message': '上傳成功', 'filename': file.filename}), 200


@excelTable.route('/uploadPdfFiles', methods=['POST'])
def upload_pdf_files():
  print("uploadPdfFiles....")

  _base_dir = current_app.config['pdfBaseDir']

  upload_type = request.form.get('uploadType')
  # 根據 uploadType 動態調整儲存目錄
  if upload_type == 'pdf1':
      _base_dir = _base_dir.replace('物料清單', '領退料單')
  if not os.path.exists(_base_dir):
    os.makedirs(_base_dir)
  print("實際儲存路徑:", _base_dir)

  files = request.files.getlist('files')  # multiple files

  if not files or len(files) == 0:
    return jsonify({'message': '沒有選擇檔案'}), 400

  saved_files = []
  for file in files:
    print("file.filename:",file.filename)

    if not file.filename.endswith('.pdf'):
      return jsonify({'message': f'檔案 {file.filename} 不是 PDF'}), 400

    filename = file.filename
    filename = os.path.basename(filename)
    file_path = os.path.join(_base_dir, filename)

    if os.path.exists(file_path):
      filename = get_unique_filename(_base_dir, filename, "copy")
      file_path = os.path.join(_base_dir, filename)

    file.save(file_path)
    saved_files.append(filename)

  return jsonify({
    'message': f'{len(saved_files)} 個 PDF 檔案上傳成功',
    'files': saved_files,
    'status': True
  })


# 內部：套用 BOM 差異(原 apply_bom_diffs 的核心；不對外成 route)
def _apply_bom_diffs_tx(s, material: Material, ops: list):
    """
    直接把差異（add/update/remove）套用到 DB。回傳(結果摘要, material 的最新 BOM 列表)。
    """
    # 先抓現有 BOM，建立索引
    existing = s.query(Bom).filter(Bom.material_id == material.id).all()
    by_key, by_mnum = {}, {}
    for b in existing:
        k = (str(b.seq_num).strip(), str(b.material_num).strip())
        by_key[k] = b
        by_mnum.setdefault(str(b.material_num).strip(), []).append(b)

    def _to_int(x, default=0):
        try:
            if x is None or (isinstance(x, str) and x.strip() == ""):
                return default
            return int(float(x))
        except Exception:
            return default

    def _recompute_non_qty(b: Bom):
        req = _to_int(b.req_qty, 0)
        picked = _to_int(b.pick_qty, 0)
        b.non_qty = max(req - picked, 0)

    results = {"added": [], "updated": [], "removed": []}

    for op in ops:
        action = (op.get("action") or "").strip().lower()
        seq_num = str(op.get("seq_num") or "").strip()
        mnum = str(op.get("material_num") or "").strip()
        if not mnum:
            continue

        target = by_key.get((seq_num, mnum))
        if not target and seq_num == "" and mnum in by_mnum and len(by_mnum[mnum]) == 1:
            target = by_mnum[mnum][0]

        if action == "add":
            comment = op.get("mtl_comment", op.get("material_comment", ""))
            qty = op.get("qty", op.get("req_qty", 0))
            start_date = op.get("start_date") or material.material_delivery_date
            receive = op.get("receive", True)

            new_bom = Bom(
                material_id=material.id,
                seq_num=seq_num or op.get("id") or "0",
                material_num=mnum,
                material_comment=str(comment or ""),
                req_qty=_to_int(qty, 0),
                pick_qty=0,
                lack_qty=0,
                non_qty=0,
                receive=bool(receive),
                start_date=str(start_date or ""),
            )
            _recompute_non_qty(new_bom)
            s.add(new_bom)
            s.flush()
            by_key[(str(new_bom.seq_num), mnum)] = new_bom
            by_mnum.setdefault(mnum, []).append(new_bom)
            results["added"].append(new_bom.get_dict())

        elif action == "update":
            if not target:
                results["updated"].append({"material_num": mnum, "seq_num": seq_num, "skipped": "not_found"})
                continue
            new_comment = op.get("mtl_comment_new", op.get("material_comment"))
            new_qty = op.get("qty_new", op.get("req_qty"))
            new_seq = op.get("seq_num_new")
            changed = False
            if new_comment is not None and str(target.material_comment) != str(new_comment):
                target.material_comment = str(new_comment); changed = True
            if new_qty is not None:
                v = _to_int(new_qty, target.req_qty or 0)
                if v != (target.req_qty or 0):
                    target.req_qty = v
                    _recompute_non_qty(target)
                    changed = True
            if new_seq is not None:
                ns = str(new_seq).strip()
                if ns and ns != str(target.seq_num):
                    old_key = (str(target.seq_num), mnum)
                    target.seq_num = ns
                    by_key.pop(old_key, None)
                    by_key[(ns, mnum)] = target
                    changed = True
            if changed:
                results["updated"].append(target.get_dict())
            else:
                results["updated"].append({"material_num": mnum, "seq_num": seq_num, "noop": True})

        elif action == "remove":
            if not target:
                candidates = by_mnum.get(mnum, [])
                if not candidates:
                    results["removed"].append({"material_num": mnum, "seq_num": seq_num, "skipped": "not_found"})
                    continue
                for b in list(candidates):
                    s.delete(b)
                    results["removed"].append({"id": b.id, "material_num": b.material_num, "seq_num": b.seq_num})
                    by_key.pop((str(b.seq_num), mnum), None)
                by_mnum.pop(mnum, None)
            else:
                s.delete(target)
                results["removed"].append({"id": target.id, "material_num": target.material_num, "seq_num": target.seq_num})
                by_key.pop((str(target.seq_num), mnum), None)
                if mnum in by_mnum:
                    by_mnum[mnum] = [b for b in by_mnum[mnum] if b.id != target.id]
                    if not by_mnum[mnum]:
                        by_mnum.pop(mnum, None)

    # 取回最新 BOM
    latest = s.query(Bom).filter(Bom.material_id == material.id).all()
    return results, [b.get_dict() for b in latest]

# 封裝：從 excel_modify 讀檔 → 產生差異清單(沿用 read_all_excel_files 的清洗規則)
def _collect_bom_diffs_from_modify_dir(s, material: Material):
    base_in = os.path.abspath(current_app.config['baseDir'])            # e.g. ...\excel_in
    modify_dir = base_in.replace("_in", "_modify")                      # ...\excel_modify
    out_dir = base_in.replace("_in", "_out")                            # ...\excel_out
    sheet_bom = current_app.config.get('excel_bom_sheet') or 1          # 系統指定索引 1 = BOM

    os.makedirs(modify_dir, exist_ok=True)
    os.makedirs(out_dir, exist_ok=True)

    order_num = normalize_order_number(material.order_num)
    files = [f for f in os.listdir(modify_dir)
        if os.path.isfile(os.path.join(modify_dir, f)) and f.endswith(('.xlsx', '.xls'))]

    modify_ops = []
    would_process_files = []

    # 先抓現有 BOM → map
    existing_bom = s.query(Bom).filter(Bom.material_id == material.id).all()
    existing_map = {str(b.material_num).strip(): b for b in existing_bom}

    for fname in files:
        path = os.path.join(modify_dir, fname)
        base_no_copy = re.sub(r'_copy_\d+', '', fname)  # 與 read_all_excel_files 同邏輯防重複

        already = s.query(exists().where(ProcessedFile.file_name == base_no_copy)).scalar()
        if already:
            print(f"[excelModifyTable] 檔案已處理，略過：{fname}")
            continue

        # 讀 BOM sheet（同 read_all_excel_files 的清洗）
        try:
            #bom_df = pd.read_excel(path, sheet_name=sheet_bom).fillna('')
            #
            bom_df = pd.read_excel(path, sheet_name=sheet_bom)
            # 文本欄位可補空字串
            for col in ['物料', '物料說明']:
              if col in bom_df.columns:
                bom_df[col] = bom_df[col].fillna('')
            # 數值欄位轉數字
            for col in ['物料短缺', '需求數量', '預留項目']:
              if col in bom_df.columns:
                bom_df[col] = pd.to_numeric(bom_df[col], errors='coerce').fillna(0).astype(int)
            #
        except Exception:
            # 某些舊檔用 openpyxl 先驗 sheet 存在
            with open(path, 'rb') as f:
                wb = openpyxl.load_workbook(filename=f, read_only=True)
                if isinstance(sheet_bom, str):
                    if sheet_bom not in wb.sheetnames:
                        print(f"[excelModifyTable] 缺少工作表：{sheet_bom}，略過 {fname}")
                        continue
                else:
                    # 如果用索引，這支分支不做；直接 raise
                    raise
            bom_df = pd.read_excel(path, sheet_name=sheet_bom).fillna('')

        if '訂單' not in bom_df.columns:
            print(f"[excelModifyTable] 缺少『訂單』欄，略過 {fname}")
            continue

        # 訂單欄正規化（比照 read_all_excel_files）
        bom_df.iloc[:, 0] = (
            bom_df.iloc[:, 0]
            .apply(normalize_order_number)
            .replace('nan', '')
            .astype(str)
        )

        rows = bom_df[bom_df['訂單'] == order_num]
        if rows.empty:
            print(f"[excelModifyTable] 檔案 {fname} 無此單號 {order_num}，略過")
            continue

        # Excel(新) → map
        incoming_map = {}
        for _, r in rows.iterrows():
            mnum = str(clean_nan(r.get('物料')) or '').strip()
            if not mnum:
                continue
            incoming_map[mnum] = {
                'seq_num':          clean_nan(r.get('預留項目')),
                'material_comment': clean_nan(r.get('物料說明')) or '',
                'req_qty':          clean_nan(r.get('需求數量')) or 0,
            }

        # remove
        for mnum, old in existing_map.items():
            if mnum not in incoming_map:
                modify_ops.append({
                    'action':      'remove',
                    'material_id': material.id,
                    'seq_num':     old.seq_num,
                    'material_num': mnum,
                })

        # add / update
        for mnum, new in incoming_map.items():
            old = existing_map.get(mnum)
            if not old:
                modify_ops.append({
                    'action':      'add',
                    'material_id': material.id,
                    'seq_num':     new['seq_num'],
                    'material_num': mnum,
                    'mtl_comment': new['material_comment'],
                    'qty':         new['req_qty'],
                    'start_date':  material.material_delivery_date,
                })
            else:
                changed = False
                old_c = str(old.material_comment or '')
                new_c = str(new['material_comment'] or '')
                if old_c != new_c:
                    changed = True
                try:
                    old_q = float(old.req_qty or 0)
                    new_q = float(new['req_qty'] or 0)
                    if old_q != new_q:
                        changed = True
                except Exception:
                    changed = True

                if changed:
                    modify_ops.append({
                        'action':          'update',
                        'material_id':     material.id,
                        'seq_num':         new['seq_num'] if new['seq_num'] is not None else old.seq_num,
                        'material_num':    mnum,
                        'mtl_comment_new': new['material_comment'],
                        'qty_new':         new['req_qty'],
                    })

        would_process_files.append((fname, path, base_no_copy, out_dir))

    return modify_ops, would_process_files


@excelTable.route('/modifyExcelFiles', methods=['POST'])
def modify_excel_files():
    print("modifyExcelFiles...")

    """
    讀取 excel_modify 內的 Excel → 產生 BOM 差異 → （預設）直接寫入 DB → 移檔到 excel_out。
    參數：
      id / material_id: 工單編號
      dry_run: true 則只比對不寫庫、不移檔（預設 False）
    回傳：
      status, message, results(added/updated/removed), bom(最新), processedFiles(實際移檔清單)
    """
    payload = request.get_json(silent=True) or {}
    material_id = payload.get("material_id") or payload.get("id")
    #id = payload.get("id")
    dry_run = bool(payload.get("dry_run", False))

    print("material_id:",material_id)

    #if not material_id or not id:
    if not material_id:
        return jsonify({"status": False, "message": "缺少 material_id / id"}), 400

    s = Session()
    try:
        material = s.query(Material).get(int(material_id))
        if not material:
            return jsonify({"status": False, "message": "找不到對應的 Material"}), 404

        # 1) 收集差異（不改 DB）
        # results:「差異」明細，讓前端知道這次到底動了哪些資料
        #   results.added: Bom[]：新增的 BOM 列表（每一筆通常是 Bom.get_dict() 的結果）。
        #   results.updated: (Bom | Meta)[]：更新過的 BOM。若某筆沒有任何變化，會回 "noop": true；若找不到對應舊資料，會回 "skipped": "not_found"。
        #   results.removed: Meta[]：被刪除的 BOM（通常包含 id / material_num / seq_num 等識別欄位）。
        modify_ops, files_to_move = _collect_bom_diffs_from_modify_dir(s, material)
        if not modify_ops:
          return jsonify({
            "status": False,
            "message": "沒有差異或沒有可處理的檔案(請上傳修正工單)",
            "results": {"added": [], "updated": [], "removed": []},
            "bom": [b.get_dict() for b in s.query(Bom).filter(Bom.material_id == material.id).all()],
            "processedFiles": []
          })

        if dry_run:
          # 僅比對，不寫庫、不移檔、不記 ProcessedFile
          return jsonify({
            "status": True,
            "message": "dry_run 僅比對完成，未寫入資料庫、未移檔",
            "diff_ops": modify_ops
          })

        # 2) 寫入 DB（apply diffs）
        results, latest_bom = _apply_bom_diffs_tx(s, material, modify_ops)

        # 3) 記錄已處理並移檔到 excel_out（尾碼 mdf）
        base_in = os.path.abspath(current_app.config['baseDir'])
        out_dir = base_in.replace("_in", "_out")
        processed = []
        for fname, path, base_no_copy, out_dir in files_to_move:
            pf = ProcessedFile(file_name=base_no_copy)
            s.add(pf)
            s.flush()

            new_name = get_unique_filename(out_dir, fname, "mdf")
            os.makedirs(out_dir, exist_ok=True)
            shutil.move(path, os.path.join(out_dir, new_name))
            processed.append(new_name)

        s.commit()

        return jsonify({
          "status": True,
          "message": "BOM 差異已套用並完成移檔",
          "results": results,
          "bom": latest_bom,
          "processedFiles": processed
        })

    except Exception as e:
        s.rollback()
        return jsonify({"status": False, "message": str(e)}), 500
    finally:
        try:
            s.close()
        except:
            pass

