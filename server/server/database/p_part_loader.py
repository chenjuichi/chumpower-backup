# server/database/p_part_loader.py

import os
from openpyxl import load_workbook
from flask import current_app

from tables import Session
from p_tables import P_Part

def import_p_part_from_excel(base_dir: str | None = None):
    """
    base_dir:
      - 若為 None：在 Flask 環境下，使用 current_app.config['baseDir']
      - 若有給值：直接用呼叫者指定的 base_dir

    從 baseDir 對應的 server 目錄底下，尋找所有 Excel 檔 (.xlsx/.xlsm/.xltx/.xltm)，
    找到工作表「配件工作中心資料表-0922 (2)」，把下列欄位寫入 p_part table：

        Excel 欄位                 ->  P_Part 欄位
        製程代號(標準內文碼)       ->  part_code
        製程說明(作業描述)         ->  part_comment
        工作中心                  ->  work_code
        工作中心名稱              ->  work_name
        成本中心                  ->  cost_code
        成本中心名稱              ->  cost_name

    並且依照「在工作表中的順序」計算 process_step_code：
        若同一張表內共有 m 筆有效製程代號：
          第一筆 -> process_step_code = m
          第二筆 -> m-1
          ...
          第 m 筆 -> 1
    """

    #_base_dir = current_app.config['baseDir']
    if base_dir is None:
        # 給 Flask 用
        _base_dir = current_app.config['baseDir']
    else:
        # 給 CLI / 測試用
        _base_dir = base_dir

    _server_dir = _base_dir.replace("excel_in", "server")

    print("import_p_part_from_excel(), _server_dir:", _server_dir)

    target_sheet_name = "配件工作中心資料表-0922 (2)"

    # 欄位名稱候選
    col_process_code_candidates = [
        "製程代號 \n(標準內文碼)",
        "製程代號(標準內文碼)",
        "製程代號 (標準內文碼)",
    ]
    col_process_comment_candidates = [
        "製程說明(作業描述)",
        "製程說明 (作業描述)",
    ]

    col_work_code_name = "工作中心"
    col_work_name_name = "工作中心名稱"
    col_cost_code_name = "成本中心"
    col_cost_name_name = "成本中心名稱"

    # openpyxl 支援的格式
    valid_exts = (".xlsx", ".xlsm", ".xltx", ".xltm")

    s = Session()

    # 先清空 p_part
    deleted_rows = s.query(P_Part).delete()
    s.commit()
    print(f"import_p_part_from_excel(): 已清空 p_part 共 {deleted_rows} 筆")

    insert_count = 0

    for root, dirs, files in os.walk(_server_dir):
        for fname in files:
            lower = fname.lower()

            # 只處理 Excel 格式
            if not lower.endswith(valid_exts) and not lower.endswith(".xls"):
                continue

            full_path = os.path.join(root, fname)

            # 舊 .xls 先略過，提醒另存新版
            if lower.endswith(".xls"):
                print(f"  檔案 {fname} 為舊版 .xls，openpyxl 不支援，請另存為 .xlsx 後再使用。")
                continue

            print(f"處理 Excel 檔案: {full_path}")

            try:
                wb = load_workbook(full_path, data_only=True)
            except Exception as e:
                print(f"  載入工作簿失敗: {e}")
                continue

            if target_sheet_name not in wb.sheetnames:
                print(f"  檔案 {fname} 沒有工作表「{target_sheet_name}」，略過。")
                continue

            ws = wb[target_sheet_name]

            # 讀第一列當表頭
            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, values_only=True),
                None
            )
            if not header_row:
                print(f"  檔案 {fname} 的表頭是空的，略過。")
                continue

            # 先找欄位 index（1-based）
            col_idx_process_code = None
            col_idx_process_comment = None
            col_idx_work_code = None
            col_idx_work_name = None
            col_idx_cost_code = None
            col_idx_cost_name = None

            header_strs = [str(h).strip() if h is not None else "" for h in header_row]

            # 製程代號欄
            for idx, val in enumerate(header_row, start=1):
                if val is None:
                    continue
                val_str = str(val).strip()
                if val_str in col_process_code_candidates:
                    col_idx_process_code = idx
                    break
            # 若沒剛好等於，就模糊找「製程代號」
            if col_idx_process_code is None:
                for idx, val in enumerate(header_row, start=1):
                    if val is None:
                        continue
                    val_str = str(val).strip()
                    if "製程代號" in val_str:
                        col_idx_process_code = idx
                        print(f"  檔案 {fname} 發現相似欄位名稱: {val_str} (col {idx})")
                        break

            # 製程說明欄
            for idx, val in enumerate(header_row, start=1):
                if val is None:
                    continue
                val_str = str(val).strip()
                if val_str in col_process_comment_candidates:
                    col_idx_process_comment = idx
                    break

            # 工作中心、工作中心名稱、成本中心、成本中心名稱
            for idx, val in enumerate(header_row, start=1):
                if val is None:
                    continue
                val_str = str(val).strip()
                if val_str == col_work_code_name:
                    col_idx_work_code = idx
                elif val_str == col_work_name_name:
                    col_idx_work_name = idx
                elif val_str == col_cost_code_name:
                    col_idx_cost_code = idx
                elif val_str == col_cost_name_name:
                    col_idx_cost_name = idx

            if col_idx_process_code is None:
                print(f"  檔案 {fname} 找不到「製程代號」欄位，略過。表頭: {header_strs}")
                continue

            def get_cell(row, idx):
                """idx 為 1-based column index，若不存在或為 None 回傳空字串"""
                if idx is None:
                    return ""
                if idx - 1 >= len(row):
                    return ""
                val = row[idx - 1]
                return "" if val is None else str(val).strip()

            # 🔹 先把所有有用的列暫存起來，之後才算 process_step_code
            rows_data = []         # 每一筆都是 dict，含 raw part_code / 其它欄位
            codes_for_mapping = [] # 只放「用來算 step 的順序 code」

            for row in ws.iter_rows(min_row=2, values_only=True):
                part_code = get_cell(row, col_idx_process_code)
                if not part_code:
                    continue  # 沒有製程代號就沒意義，略過

                part_comment = get_cell(row, col_idx_process_comment)
                work_code = get_cell(row, col_idx_work_code)
                work_name = get_cell(row, col_idx_work_name)
                cost_code = get_cell(row, col_idx_cost_code)
                cost_name = get_cell(row, col_idx_cost_name)

                # key_code 是拿來算順序用的 code
                #（如果你希望不去掉 'B'，就改成 key_code = part_code）
                key_code = part_code[1:] if part_code.startswith("B") else part_code

                rows_data.append({
                    "part_code": part_code,
                    "part_comment": part_comment,
                    "work_code": work_code,
                    "work_name": work_name,
                    "cost_code": cost_code,
                    "cost_name": cost_name,
                    "key_code": key_code,
                })
                codes_for_mapping.append(key_code)

            if not rows_data:
                print(f"  檔案 {fname} 這個工作表沒有有效製程代號資料，略過。")
                continue

            # 🔹 建立「key_code -> process_step_code」的 mapping
            m = len(codes_for_mapping)
            print(f"  檔案 {fname} 有 {m} 筆製程代號，要寫入 process_step_code")

            key_to_step = {}
            for idx, key_code in enumerate(codes_for_mapping, start=1):
                # 第一筆 -> m, 第二筆 -> m-1, ..., 第 m 筆 -> 1
                value = m - idx + 1
                key_to_step[key_code] = value

            # 🔹 依序建立 P_Part，每一筆帶上對應的 process_step_code
            for item in rows_data:
                step = key_to_step.get(item["key_code"], 0)

                part = P_Part(
                    #process_step_code=step ,
                    process_step_code = step + 1000,
                    part_code=item["part_code"],
                    part_comment=item["part_comment"],
                    work_code=item["work_code"],
                    work_name=item["work_name"],
                    cost_code=item["cost_code"],
                    cost_name=item["cost_name"],
                )
                s.add(part)
                insert_count += 1

            s.commit()

    s.close()
    print(f"import_p_part_from_excel(): 匯入完成，共新增 {insert_count} 筆 p_part")
    #return insert_count


if __name__ == "__main__":
    # 這裡自己給一個 base_dir（模擬 Flask 的 config['baseDir']）
    base_dir = r"C:\vue\chumpower\excel_in"

    # 如果要確保 DB / Table 已建立，可以在這裡 import BASE, engine
    from tables import BASE, engine
    # BASE.metadata.create_all(engine)  # 如果需要才打開

    import_p_part_from_excel(base_dir=base_dir)