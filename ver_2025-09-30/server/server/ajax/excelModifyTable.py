import os, re, shutil
import pandas as pd
import openpyxl
from sqlalchemy import exists

import psutil
from sqlalchemy import exists

from openpyxl import Workbook, load_workbook
from database.tables import Session, Material, Bom, ProcessedFile
from flask import Blueprint, jsonify, request, current_app

import warnings
warnings.filterwarnings("ignore", category=FutureWarning,
                        message=".*Setting an item of incompatible dtype.*")

excelModifyTable = Blueprint('excelModifyTable', __name__)

from log_util import setup_logger
logger = setup_logger(__name__)  # 每個模組用自己的名稱


# ------------------------------------------------------------------


def get_unique_filename(target_dir, filename, chip):
    base, ext = os.path.splitext(filename)  # 分離檔案名稱與副檔名
    counter = 1
    unique_filename = filename
    while os.path.exists(os.path.join(target_dir, unique_filename)):  # 檢查檔案是否已存在
        unique_filename = f"{base}_{chip}_{counter}{ext}"  # 為檔名新增後綴
        counter += 1
    return unique_filename

def convert_date(date_value):
    try:
        return pd.to_datetime(date_value).date()  # 使用 pandas 轉換為日期
    except (ValueError, TypeError):
        return None

def close_open_file(filepath):
    # 遍歷所有進程
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # 檢查該進程打開的所有文件
            for item in proc.open_files():
                if item.path == filepath:
                    # 如果文件被打開，嘗試關閉
                    print(f"關閉進程 {proc.pid} 打開的文件: {filepath}")
                    proc.terminate()  # 終止進程
                    proc.wait()  # 等待進程終止
                    return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            continue
    return False

def pad_zeros(str):
    return str.zfill(4)

def clean_nan(value):
    """清理單一值的 NaN / NaT，轉為 None"""
    if pd.isna(value) or str(value).lower() == 'nan':
        return None
    return value


def normalize_order_number(value):
    """
    將單號轉為純字串，不帶小數點，並處理 NaN。
    例如：121100017702.0 -> '121100017702'
    """
    if pd.isna(value) or str(value).lower() == 'nan':
        return ''
    try:
        return str(int(float(value)))
    except (ValueError, TypeError):
        return str(value).strip()


# 內部：套用 BOM 差異(原 apply_bom_diffs 的核心；不對外成 route)
def _apply_bom_diffs_tx(s, material: Material, ops: list):
    """
    直接把差異（add/update/remove）套用到 DB。回傳(結果摘要, material 的最新 BOM 列表)。
    """
    # 先抓現有 BOM，建立索引
    existing = s.query(Bom).filter(Bom.material_id == material.id).all()
    by_key, by_mnum = {}, {}
    for b in existing:
        k = (str(b.seq_num).strip(), str(b.material_num).strip())
        by_key[k] = b
        by_mnum.setdefault(str(b.material_num).strip(), []).append(b)

    def _to_int(x, default=0):
        try:
            if x is None or (isinstance(x, str) and x.strip() == ""):
                return default
            return int(float(x))
        except Exception:
            return default

    def _recompute_non_qty(b: Bom):
        req = _to_int(b.req_qty, 0)
        picked = _to_int(b.pick_qty, 0)
        b.non_qty = max(req - picked, 0)

    results = {"added": [], "updated": [], "removed": []}

    for op in ops:
        action = (op.get("action") or "").strip().lower()
        seq_num = str(op.get("seq_num") or "").strip()
        mnum = str(op.get("material_num") or "").strip()
        if not mnum:
            continue

        target = by_key.get((seq_num, mnum))
        if not target and seq_num == "" and mnum in by_mnum and len(by_mnum[mnum]) == 1:
            target = by_mnum[mnum][0]

        if action == "add":
            comment = op.get("mtl_comment", op.get("material_comment", ""))
            qty = op.get("qty", op.get("req_qty", 0))
            start_date = op.get("start_date") or material.material_delivery_date
            receive = op.get("receive", True)

            new_bom = Bom(
                material_id=material.id,
                seq_num=seq_num or op.get("id") or "0",
                material_num=mnum,
                material_comment=str(comment or ""),
                req_qty=_to_int(qty, 0),
                pick_qty=0,
                lack_qty=0,
                non_qty=0,
                receive=bool(receive),
                start_date=str(start_date or ""),
            )
            _recompute_non_qty(new_bom)
            s.add(new_bom)
            s.flush()
            by_key[(str(new_bom.seq_num), mnum)] = new_bom
            by_mnum.setdefault(mnum, []).append(new_bom)
            results["added"].append(new_bom.get_dict())

        elif action == "update":
            if not target:
                results["updated"].append({"material_num": mnum, "seq_num": seq_num, "skipped": "not_found"})
                continue
            new_comment = op.get("mtl_comment_new", op.get("material_comment"))
            new_qty = op.get("qty_new", op.get("req_qty"))
            new_seq = op.get("seq_num_new")
            changed = False
            if new_comment is not None and str(target.material_comment) != str(new_comment):
                target.material_comment = str(new_comment); changed = True
            if new_qty is not None:
                v = _to_int(new_qty, target.req_qty or 0)
                if v != (target.req_qty or 0):
                    target.req_qty = v
                    _recompute_non_qty(target)
                    changed = True
            if new_seq is not None:
                ns = str(new_seq).strip()
                if ns and ns != str(target.seq_num):
                    old_key = (str(target.seq_num), mnum)
                    target.seq_num = ns
                    by_key.pop(old_key, None)
                    by_key[(ns, mnum)] = target
                    changed = True
            if changed:
                results["updated"].append(target.get_dict())
            else:
                results["updated"].append({"material_num": mnum, "seq_num": seq_num, "noop": True})

        elif action == "remove":
            if not target:
                candidates = by_mnum.get(mnum, [])
                if not candidates:
                    results["removed"].append({"material_num": mnum, "seq_num": seq_num, "skipped": "not_found"})
                    continue
                for b in list(candidates):
                    s.delete(b)
                    results["removed"].append({"id": b.id, "material_num": b.material_num, "seq_num": b.seq_num})
                    by_key.pop((str(b.seq_num), mnum), None)
                by_mnum.pop(mnum, None)
            else:
                s.delete(target)
                results["removed"].append({"id": target.id, "material_num": target.material_num, "seq_num": target.seq_num})
                by_key.pop((str(target.seq_num), mnum), None)
                if mnum in by_mnum:
                    by_mnum[mnum] = [b for b in by_mnum[mnum] if b.id != target.id]
                    if not by_mnum[mnum]:
                        by_mnum.pop(mnum, None)

    # 取回最新 BOM
    latest = s.query(Bom).filter(Bom.material_id == material.id).all()
    return results, [b.get_dict() for b in latest]

# 封裝：從 excel_modify 讀檔 → 產生差異清單(沿用 read_all_excel_files 的清洗規則)
def _collect_bom_diffs_from_modify_dir(s, material: Material):
    base_in = os.path.abspath(current_app.config['baseDir'])            # e.g. ...\excel_in
    modify_dir = base_in.replace("_in", "_modify")                      # ...\excel_modify
    out_dir = base_in.replace("_in", "_out")                            # ...\excel_out
    sheet_bom = current_app.config.get('excel_bom_sheet') or 1          # 系統指定索引 1 = BOM

    os.makedirs(modify_dir, exist_ok=True)
    os.makedirs(out_dir, exist_ok=True)

    order_num = normalize_order_number(material.order_num)
    files = [f for f in os.listdir(modify_dir)
        if os.path.isfile(os.path.join(modify_dir, f)) and f.endswith(('.xlsx', '.xls'))]

    modify_ops = []
    would_process_files = []

    # 先抓現有 BOM → map
    existing_bom = s.query(Bom).filter(Bom.material_id == material.id).all()
    existing_map = {str(b.material_num).strip(): b for b in existing_bom}

    for fname in files:
        path = os.path.join(modify_dir, fname)
        base_no_copy = re.sub(r'_copy_\d+', '', fname)  # 與 read_all_excel_files 同邏輯防重複

        already = s.query(exists().where(ProcessedFile.file_name == base_no_copy)).scalar()
        if already:
            print(f"[excelModifyTable] 檔案已處理，略過：{fname}")
            continue

        # 讀 BOM sheet（同 read_all_excel_files 的清洗）
        try:
            #bom_df = pd.read_excel(path, sheet_name=sheet_bom).fillna('')
            #
            bom_df = pd.read_excel(path, sheet_name=sheet_bom)
            # 文本欄位可補空字串
            for col in ['物料', '物料說明']:
              if col in bom_df.columns:
                bom_df[col] = bom_df[col].fillna('')
            # 數值欄位轉數字
            for col in ['物料短缺', '需求數量', '預留項目']:
              if col in bom_df.columns:
                bom_df[col] = pd.to_numeric(bom_df[col], errors='coerce').fillna(0).astype(int)
            #
        except Exception:
            # 某些舊檔用 openpyxl 先驗 sheet 存在
            with open(path, 'rb') as f:
                wb = openpyxl.load_workbook(filename=f, read_only=True)
                if isinstance(sheet_bom, str):
                    if sheet_bom not in wb.sheetnames:
                        print(f"[excelModifyTable] 缺少工作表：{sheet_bom}，略過 {fname}")
                        continue
                else:
                    # 如果用索引，這支分支不做；直接 raise
                    raise
            bom_df = pd.read_excel(path, sheet_name=sheet_bom).fillna('')

        if '訂單' not in bom_df.columns:
            print(f"[excelModifyTable] 缺少『訂單』欄，略過 {fname}")
            continue

        # 訂單欄正規化（比照 read_all_excel_files）
        bom_df.iloc[:, 0] = (
            bom_df.iloc[:, 0]
            .apply(normalize_order_number)
            .replace('nan', '')
            .astype(str)
        )

        rows = bom_df[bom_df['訂單'] == order_num]
        if rows.empty:
            print(f"[excelModifyTable] 檔案 {fname} 無此單號 {order_num}，略過")
            continue

        # Excel(新) → map
        incoming_map = {}
        for _, r in rows.iterrows():
            mnum = str(clean_nan(r.get('物料')) or '').strip()
            if not mnum:
                continue
            incoming_map[mnum] = {
                'seq_num':          clean_nan(r.get('預留項目')),
                'material_comment': clean_nan(r.get('物料說明')) or '',
                'req_qty':          clean_nan(r.get('需求數量')) or 0,
            }

        # remove
        for mnum, old in existing_map.items():
            if mnum not in incoming_map:
                modify_ops.append({
                    'action':      'remove',
                    'material_id': material.id,
                    'seq_num':     old.seq_num,
                    'material_num': mnum,
                })

        # add / update
        for mnum, new in incoming_map.items():
            old = existing_map.get(mnum)
            if not old:
                modify_ops.append({
                    'action':      'add',
                    'material_id': material.id,
                    'seq_num':     new['seq_num'],
                    'material_num': mnum,
                    'mtl_comment': new['material_comment'],
                    'qty':         new['req_qty'],
                    'start_date':  material.material_delivery_date,
                })
            else:
                changed = False
                old_c = str(old.material_comment or '')
                new_c = str(new['material_comment'] or '')
                if old_c != new_c:
                    changed = True
                try:
                    old_q = float(old.req_qty or 0)
                    new_q = float(new['req_qty'] or 0)
                    if old_q != new_q:
                        changed = True
                except Exception:
                    changed = True

                if changed:
                    modify_ops.append({
                        'action':          'update',
                        'material_id':     material.id,
                        'seq_num':         new['seq_num'] if new['seq_num'] is not None else old.seq_num,
                        'material_num':    mnum,
                        'mtl_comment_new': new['material_comment'],
                        'qty_new':         new['req_qty'],
                    })

        would_process_files.append((fname, path, base_no_copy, out_dir))

    return modify_ops, would_process_files


@excelModifyTable.route("/modifyExcelFiles", methods=["GET"])
def modify_excel_files():
    print("modifyExcelFiles...")

    """
    讀取 excel_modify 內的 Excel → 產生 BOM 差異 → （預設）直接寫入 DB → 移檔到 excel_out。
    參數：
      id / material_id: 工單編號
      dry_run: true 則只比對不寫庫、不移檔（預設 False）
    回傳：
      status, message, results(added/updated/removed), bom(最新), processedFiles(實際移檔清單)
    """
    payload = request.get_json(silent=True) or {}
    material_id = payload.get("material_id") or payload.get("id")
    #id = payload.get("id")
    dry_run = bool(payload.get("dry_run", False))

    print("material_id:",material_id)

    #if not material_id or not id:
    if not material_id:
        return jsonify({"status": False, "message": "缺少 material_id / id"}), 400

    s = Session()
    try:
        material = s.query(Material).get(int(material_id))
        if not material:
            return jsonify({"status": False, "message": "找不到對應的 Material"}), 404

        # 1) 收集差異（不改 DB）
        # results:「差異」明細，讓前端知道這次到底動了哪些資料
        #   results.added: Bom[]：新增的 BOM 列表（每一筆通常是 Bom.get_dict() 的結果）。
        #   results.updated: (Bom | Meta)[]：更新過的 BOM。若某筆沒有任何變化，會回 "noop": true；若找不到對應舊資料，會回 "skipped": "not_found"。
        #   results.removed: Meta[]：被刪除的 BOM（通常包含 id / material_num / seq_num 等識別欄位）。
        modify_ops, files_to_move = _collect_bom_diffs_from_modify_dir(s, material)
        if not modify_ops:
          return jsonify({
            "status": True,
            "message": "沒有差異或沒有可處理的檔案",
            "results": {"added": [], "updated": [], "removed": []},
            "bom": [b.get_dict() for b in s.query(Bom).filter(Bom.material_id == material.id).all()],
            "processedFiles": []
          })

        if dry_run:
          # 僅比對，不寫庫、不移檔、不記 ProcessedFile
          return jsonify({
            "status": True,
            "message": "dry_run 僅比對完成，未寫入資料庫、未移檔",
            "diff_ops": modify_ops
          })

        # 2) 寫入 DB（apply diffs）
        results, latest_bom = _apply_bom_diffs_tx(s, material, modify_ops)

        # 3) 記錄已處理並移檔到 excel_out（尾碼 mdf）
        base_in = os.path.abspath(current_app.config['baseDir'])
        out_dir = base_in.replace("_in", "_out")
        processed = []
        for fname, path, base_no_copy, out_dir in files_to_move:
            pf = ProcessedFile(file_name=base_no_copy)
            s.add(pf)
            s.flush()

            new_name = get_unique_filename(out_dir, fname, "mdf")
            os.makedirs(out_dir, exist_ok=True)
            shutil.move(path, os.path.join(out_dir, new_name))
            processed.append(new_name)

        s.commit()

        return jsonify({
          "status": True,
          "message": "BOM 差異已套用並完成移檔",
          "results": results,
          "bom": latest_bom,
          "processedFiles": processed
        })

    except Exception as e:
        s.rollback()
        return jsonify({"status": False, "message": str(e)}), 500
    finally:
        try:
            s.close()
        except:
            pass


'''
@excelModifyTable.route("/modifyExcelFiles", methods=['POST'])
def modify_excel_files():

  data = request.json
  material_id = data.get("material_id")
  _id = data.get("id")
  _material_id = int(material_id)
  print("_id, _material_id:",_id, _material_id, type(_id), type(_material_id))

  global global_var

  return_value = False
  return_message1 = '錯誤, excel檔案內沒有該筆工單!'
  return_message = ''
  modifyBom = []  # 用於儲存 BOM 資料的清單

  _base_dir = current_app.config['baseDir']
  _modify_dir = _base_dir.replace("_in", "_modify")
  _target_dir = _base_dir.replace("_in", "_out")
  print("read modify excel files, 目錄: ", _modify_dir)
  print("move excel files to, 目錄: ", _target_dir)
  # 讀取指定目錄下的所有指定檔案名稱
  #files = [f for f in os.listdir(_modify_dir) if os.path.isfile(os.path.join(_modify_dir, f)) and f.startswith('Report_') and f.endswith('.xlsx')]
  files = [f for f in os.listdir(_modify_dir) if os.path.isfile(os.path.join(_modify_dir, f)) and f.endswith('.xlsx')]
  if (files):   #有工單檔案
    sheet_names_to_check = [
      current_app.config['excel_product_sheet'],
      current_app.config['excel_bom_sheet'],
    ]
    _startRow = int(current_app.config['startRow'])

    s = Session()

    find_material = s.query(Material).filter_by(id = _id).first() #找出一筆工單資料

    # 取得現有的 BOM 資料
    existing_bom = s.query(Bom).filter_by(material_id = find_material.id).all()
    existing_materials = {bom.material_num for bom in existing_bom}

    for _file_name in files:  #檔案讀取
      _path = _modify_dir + '\\' + _file_name
      global_var = _path + ' 檔案讀取中...'

      with open(_path, 'rb') as file:
        workbook = openpyxl.load_workbook(filename=file, read_only=True)
        return_value = True
        return_message1 = ''
        missing_sheets = [sheet for sheet in sheet_names_to_check if sheet not in workbook.sheetnames]

        if missing_sheets:
          return_value = False
          return_message1 = '錯誤, 工單檔案內沒有相關工作表!'
          print(return_message1)
          break

        print(sheet_names_to_check[0] + ' sheet exists, data reading...')

        #bom_df = pd.read_excel(_path, sheet_name=1)               # Second sheet for Bom
        bom_df = pd.read_excel(_path, sheet_name=1).replace({pd.NaT: None})   #為 bom_entries 清理 NaN
        bom_df = bom_df.fillna('')                                            #DataFrame 讀取後 .fillna('') 處理空值

        bom_df['訂單'] = bom_df['訂單'].fillna(0).astype(int)     # 檢查是否存在 NaN 值
        bom_df['訂單'] = bom_df['訂單'].fillna('').astype(str)    # 保持字串型態
        bom_entries = bom_df[bom_df['訂單'] == material_id]       # 查詢對應的 BOM 項目
        print(f"bom_entries 中的資料筆數: {len(bom_entries)}")

        # Insert corresponding BOM entries
        bom_data= []
        for bom_index, bom_row in bom_entries.iterrows():
          bom_data.append({
            'material_id': find_material.id,
            'seq_num': bom_row['預留項目'],
            'material_num': bom_row['物料'],
            'material_comment': bom_row['物料說明'],
            'req_qty': bom_row['需求數量'],
          })


        # 將現有的 BOM 資料加入 modifyBom
        id = 0
        for existing in existing_bom:
          id +=1
          modifyBom.append({
            "id": id,
            "material_id": existing.material_id,
            "seq_num": existing.seq_num,
            "material_num": existing.material_num,
            "mtl_comment": existing.material_comment,
            "qty": existing.req_qty,
            "start_date": existing.start_date,
            "date_alarm": '',
          })

        # 新增的 material_num
        for bom_entry in bom_data:
          material_num = bom_entry.get("material_num")
          if material_num not in existing_materials:
            id +=1
            modifyBom.append({
              "id": id,
              "material_id": find_material.id,
              "seq_num": id,
              "material_num": material_num,
              "mtl_comment": bom_entry.get("material_comment"),
              "qty": bom_entry.get("req_qty"),
              "start_date": find_material.material_delivery_date,
              "date_alarm": '',
            })
    #end for loop

    # 提交到資料庫
    #s.commit()
    #s.close()
  # end if

  return jsonify({
    'status': return_value,
    'message': return_message1,
    'modifyBom': modifyBom,
    'modifyFileName': _file_name,
  })
'''
