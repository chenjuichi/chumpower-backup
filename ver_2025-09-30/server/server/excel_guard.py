# -*- coding: utf-8 -*-
"""
excel_guard.py
整合版：
- 預設掃描 D:\chumpower_data
- 可直接指定檔名（相對於 base-dir 或絕對路徑）
- 可互動挑檔（檔案選擇器 / 文字清單）
- 輸出違規報表到與來源檔相同資料夾（自動加時間戳）

規則：
1)（主規則）若第 1 列的某欄標題為「空白」（公式或值任一為空即視為空），
   則該欄第 2 列起不得有任何值或公式
2)（可關閉）若某列第 1 欄為「空白」，則該列其他欄不得有任何值或公式
"""
import argparse
import sys
from pathlib import Path
from datetime import datetime
from typing import List, Optional

# 依賴
import pandas as pd
from openpyxl import load_workbook

DEFAULT_BASE = Path(r"D:\chumpower_data")

def is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False

def pick_file_interactive(start_dir: Path) -> Optional[Path]:
    """嘗試用檔案選擇器，失敗則回傳 None。"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            initialdir=str(start_dir),
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        root.destroy()
        return Path(path) if path else None
    except Exception:
        return None

def pick_file_from_console(candidates: List[Path]) -> Optional[Path]:
    print("\n可選擇的檔案（依修改時間新→舊）：")
    for i, p in enumerate(candidates, 1):
        ts = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        print(f"{i:2d}. {p.name:50s}  {ts}")
    try:
        raw = input("\n輸入要選擇的編號（Enter 取消）：").strip()
        if not raw:
            return None
        idx = int(raw)
    except ValueError:
        return None
    if 1 <= idx <= len(candidates):
        return candidates[idx - 1]
    return None

def resolve_input_file(args) -> Path:
    base_dir = Path(args.base_dir) if args.base_dir else DEFAULT_BASE
    base_dir.mkdir(parents=True, exist_ok=True)

    # 1) 明確指定檔名
    if args.file:
        p = Path(args.file)
        if not p.is_absolute():
            p = base_dir / p
        if not p.exists():
            print(f"找不到檔案：{p}")
            sys.exit(2)
        return p

    # 2) 互動挑檔
    if args.interactive:
        p = pick_file_interactive(base_dir)
        if p and p.exists():
            return p
        # GUI 失敗 → 文字清單
        candidates = sorted([x for x in base_dir.glob("*.xlsx") if not x.name.startswith("~")],
                            key=lambda x: x.stat().st_mtime, reverse=True)
        if not candidates:
            print(f"資料夾內沒有 .xlsx：{base_dir}")
            sys.exit(2)
        p = pick_file_from_console(candidates)
        if p:
            return p
        print("未選擇任何檔案。")
        sys.exit(2)

    # 3) 取最新
    if args.latest:
        candidates = sorted([x for x in base_dir.glob("*.xlsx") if not x.name.startswith("~")],
                            key=lambda x: x.stat().st_mtime, reverse=True)
        if not candidates:
            print(f"資料夾內沒有 .xlsx：{base_dir}")
            sys.exit(2)
        return candidates[0]

    print("請用 --file 指定檔案，或加上 --interactive / --latest")
    sys.exit(2)

def check_workbook(xlsx_path: Path, sheets: Optional[List[str]], start_col: int, start_row: int, enable_row_rule: bool):
    # 兩次載入：一次讀公式字串（data_only=False）、一次讀計算值（data_only=True）
    wb_f = load_workbook(xlsx_path, data_only=False, read_only=False)
    wb_v = load_workbook(xlsx_path, data_only=True,  read_only=True)

    report_rows = []

    target_sheets = sheets if sheets else wb_f.sheetnames
    for sheet_name in target_sheets:
        if sheet_name not in wb_f.sheetnames:
            print(f"略過不存在工作表：{sheet_name}")
            continue
        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]

        max_row = ws_f.max_row or 0
        max_col = ws_f.max_column or 0

        # === 規則 1：標題空白 → 整欄必須空（含值/公式）
        for c in range(start_col, max_col + 1):
            head_f_val = ws_f.cell(row=1, column=c).value
            head_v_val = ws_v.cell(row=1, column=c).value
            header_is_blank = is_blank(head_f_val) or is_blank(head_v_val)
            if not header_is_blank:
                continue

            for r in range(start_row, max_row + 1):
                cell_f = ws_f.cell(row=r, column=c)
                cell_v = ws_v.cell(row=r, column=c)

                val_f = cell_f.value
                val_v = cell_v.value
                has_formula = isinstance(val_f, str) and val_f.startswith("=")
                has_value = (not is_blank(val_v)) or (not is_blank(val_f))
                if has_formula or has_value:
                    report_rows.append({
                        "Sheet": sheet_name,
                        "Rule": "Blank header in row 1 → column must be empty",
                        "Cell": cell_f.coordinate,
                        "Header(1,{})".format(c): "",
                        "CachedValue": val_v,
                        "Literal/Formula": val_f if val_f is not None else ""
                    })

        # === 規則 2（可關閉）：列首空白 → 整列必須空
        if enable_row_rule:
            for r in range(start_row, max_row + 1):
                first_f_val = ws_f.cell(row=r, column=1).value
                first_v_val = ws_v.cell(row=r, column=1).value
                first_blank = is_blank(first_f_val) and is_blank(first_v_val)
                if not first_blank:
                    continue

                for c in range(start_col, max_col + 1):
                    cell_f = ws_f.cell(row=r, column=c)
                    cell_v = ws_v.cell(row=r, column=c)
                    valf = cell_f.value
                    has_formula = isinstance(valf, str) and valf.startswith("=")
                    has_value = (not is_blank(cell_v.value)) or (not is_blank(valf))
                    if has_formula or has_value:
                        report_rows.append({
                            "Sheet": sheet_name,
                            "Rule": "Blank first cell in row → row must be empty",
                            "Cell": cell_f.coordinate,
                            "HeaderA1": ws_v.cell(row=1, column=1).value,
                            "CachedValue": cell_v.value,
                            "Literal/Formula": valf if valf is not None else ""
                        })

    df = pd.DataFrame(report_rows)
    return df

def save_report(df: pd.DataFrame, xlsx_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = xlsx_path.with_name(f"{xlsx_path.stem}_violations_{ts}.xlsx")
    with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
        if df is None or df.empty:
            pd.DataFrame([{"Info": "No violations found."}]).to_excel(writer, index=False, sheet_name="Report")
        else:
            df.to_excel(writer, index=False, sheet_name="Violations")
    return out_path

def main():
    parser = argparse.ArgumentParser(description="Excel 欄位/列結構檢查")
    parser.add_argument("--base-dir", help="預設資料夾（預設：D:\\chumpower_data）")
    parser.add_argument("--file", help="檔名（可相對於 base-dir 或絕對路徑）")
    parser.add_argument("--interactive", action="store_true", help="互動挑檔（檔案選擇器；失敗則文字清單）")
    parser.add_argument("--latest", action="store_true", help="從 base-dir 取最新 .xlsx")
    parser.add_argument("--sheets", nargs="*", help="只檢查指定工作表（空白=全部）")
    parser.add_argument("--start-col", type=int, default=2, help="從第幾欄開始檢查（標題規則；預設 2）")
    parser.add_argument("--start-row", type=int, default=2, help="從第幾列開始檢查（兩規則；預設 2）")
    parser.add_argument("--no-row-rule", action="store_true", help="關閉『列首空白→整列必須空』規則")
    args = parser.parse_args()

    xlsx_path = resolve_input_file(args)
    print(f"檢查檔案：{xlsx_path}")

    df = check_workbook(
        xlsx_path=xlsx_path,
        sheets=args.sheets,
        start_col=args.start_col,
        start_row=args.start_row,
        enable_row_rule=not args.no_row_rule
    )

    out_path = save_report(df, xlsx_path)
    print(f"報表已輸出：{out_path}")
    if df is None or df.empty:
        print("檢查完成：未發現違規。")
    else:
        print(f"檢查完成：共 {len(df)} 筆違規。")

if __name__ == "__main__":
    main()
